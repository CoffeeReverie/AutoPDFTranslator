from __future__ import annotations

import html
import base64
import io
import json
import logging
import math
import os
import re
import tempfile
import time
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components

from autopdftranslator import APP_NAME, APP_VERSION, LayoutConfig, TranslationCancelled, analyze_pdf, translate_pdf

try:
    import requests
except Exception:  # pragma: no cover
    requests = None


LANGUAGE_OPTIONS = [
    ("Chinese (Simplified)", "zh-CN"),
    ("Chinese (Traditional)", "zh-TW"),
    ("Japanese", "ja"),
    ("Korean", "ko"),
    ("French", "fr"),
    ("German", "de"),
    ("Spanish", "es"),
    ("English", "en"),
]
SOURCE_LANGUAGE_OPTIONS = [("Auto Detect", "auto")] + LANGUAGE_OPTIONS

MODE_HINTS = {
    "text extraction only": "Fastest. Uses only typed text from the PDF layer.",
    "vision only": "纯视觉识别建议使用API Providers",
    "hybrid": "Uses text extraction first, then vision fallback when needed.",
}

MODE_HISTORY_LABEL = {
    "text extraction only": "Text",
    "vision only": "Vision",
    "hybrid": "Hybrid",
}

BLUEBEAM_ANNOTATION_FONT_SIZES: list[float] = [float(size) for size in range(2, 73)]
ADOBE_ANNOTATION_FONT_SIZES = BLUEBEAM_ANNOTATION_FONT_SIZES
DEFAULT_ANNOTATION_FONT_SIZE = 10.0

PLACEMENT_OPTIONS = {
    "top-left stacked": "top-left-stacked",
    "top-right stacked": "top-right-stacked",
}

API_PROVIDER_PRESETS: dict[str, dict[str, str]] = {
    "OpenAI": {
        "base_url": "https://api.openai.com/v1",
        "model": "gpt-4.1-mini",
        "provider_impl": "openai-compatible",
    },
    "Moonshot": {
        "base_url": "https://api.moonshot.cn/v1",
        "model": "moonshot-v1-8k",
        "provider_impl": "openai-compatible",
    },
    "Gemini (OpenAI-Compatible)": {
        "base_url": "https://generativelanguage.googleapis.com/v1beta/openai",
        "model": "gemini-2.0-flash",
        "provider_impl": "openai-compatible",
    },
    "DeepSeek": {
        "base_url": "https://api.deepseek.com/v1",
        "model": "deepseek-chat",
        "provider_impl": "openai-compatible",
    },
    "MiniMax": {
        "base_url": "https://api.minimax.io/v1",
        "model": "MiniMax-M2.7",
        "provider_impl": "openai-compatible",
    },
    "MiniMax (CN)": {
        "base_url": "https://api.minimaxi.com/v1",
        "model": "MiniMax-M2.7",
        "provider_impl": "openai-compatible",
    },
    "Qwen Compatible": {
        "base_url": "https://dashscope.aliyuncs.com/compatible-mode/v1",
        "model": "qwen-plus",
        "provider_impl": "openai-compatible",
    },
}

LOCAL_MODEL_PRESETS: dict[str, dict[str, str]] = {
    "Ollama": {
        "base_url": "http://localhost:11434/v1",
        "model": "translategemma:latest",
        "provider_impl": "openai-compatible",
    },
    "LM Studio": {
        "base_url": "http://localhost:1234/v1",
        "model": "local-model",
        "provider_impl": "openai-compatible",
    },
    "vLLM": {
        "base_url": "http://localhost:8000/v1",
        "model": "local-model",
        "provider_impl": "openai-compatible",
    },
    "Mock (offline)": {
        "base_url": "",
        "model": "",
        "provider_impl": "mock",
    },
}

def _app_storage_root() -> Path:
    raw = os.getenv("AUTOPDFTRANSLATOR_STORAGE_DIR", "").strip()
    root = Path(raw).expanduser() if raw else Path(__file__).resolve().parent
    try:
        root.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return root


APP_STORAGE_ROOT = _app_storage_root()
UI_CONFIG_PATH = APP_STORAGE_ROOT / ".autopdftranslator_ui_config.json"
TRANSLATION_HISTORY_PATH = APP_STORAGE_ROOT / ".autopdftranslator_history.json"
TRANSLATION_CORPUS_PATH = APP_STORAGE_ROOT / ".autopdftranslator_corpus.json"
TRANSLATION_MEMORY_PATH = APP_STORAGE_ROOT / ".autopdftranslator_memory.json"
RUNTIME_MEMORY_PATH = APP_STORAGE_ROOT / ".autopdftranslator_runtime_memory.json"
DEFAULT_USD_TO_CNY_RATE = 7.20
MAX_CORPUS_ENTRY_CHARS = 120000
MAX_MEMORY_EXAMPLE_CHARS = 600
CORPUS_UPLOAD_TYPES = ["pdf", "docx", "pptx", "xlsx", "txt", "md", "csv", "doc", "ppt", "xls"]


class _StreamlitLogHandler(logging.Handler):
    def __init__(self) -> None:
        super().__init__()
        self.lines: list[str] = []

    def emit(self, record: logging.LogRecord) -> None:
        self.lines.append(self.format(record))


def _inject_styles() -> None:
    st.markdown(
        """
<style>
    .stApp {
        background:
            radial-gradient(circle at 0% 0%, rgba(57, 117, 234, 0.10), transparent 42%),
            radial-gradient(circle at 100% 0%, rgba(57, 192, 169, 0.08), transparent 40%),
            #f6f8fc;
    }
    .hero {
        background: linear-gradient(120deg, #0f172a 0%, #1e3a8a 45%, #0f766e 100%);
        border-radius: 18px;
        padding: 28px 30px;
        margin-bottom: 16px;
        color: #ffffff;
        box-shadow: 0 14px 30px rgba(15, 23, 42, 0.24);
    }
    .hero h1 {
        margin: 0;
        font-size: 2rem;
        line-height: 1.1;
    }
    .hero p {
        margin: 8px 0 0;
        color: #dbeafe;
    }
    div[data-testid="stButton"] button[kind="primary"] {
        border-radius: 10px;
        border: 0;
        font-weight: 600;
        height: 2.9rem;
        background: linear-gradient(90deg, #2563eb, #0ea5a4);
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        filter: brightness(1.04);
    }
    .st-key-stop_translation_btn button {
        border-radius: 10px;
        font-weight: 700;
        height: 2.9rem;
    }
    .st-key-stop_translation_btn button:not(:disabled) {
        background: #dc2626 !important;
        border-color: #dc2626 !important;
        color: #ffffff !important;
    }
    .st-key-stop_translation_btn button:not(:disabled):hover {
        background: #b91c1c !important;
        border-color: #b91c1c !important;
        color: #ffffff !important;
    }
    .st-key-stop_translation_btn button:disabled {
        background: #e5e7eb !important;
        border-color: #d1d5db !important;
        color: #6b7280 !important;
    }
    div[data-testid="stProgress"] div[role="progressbar"] > div {
        background-color: #16a34a !important;
    }
    .stop-button-spacer {
        height: 70px;
    }
    .stage-progress {
        margin: 2px 0 18px;
        padding: 14px 16px 14px;
        background: rgba(255, 255, 255, 0.72);
        border: 1px solid rgba(148, 163, 184, 0.26);
        border-radius: 12px;
    }
    .stage-progress__meta {
        margin-bottom: 10px;
        color: #334155;
        font-size: 0.88rem;
        line-height: 1.35;
    }
    .stage-progress__row {
        display: flex;
        align-items: center;
        gap: 18px;
    }
    .stage-progress__main {
        flex: 1 1 auto;
        min-width: 180px;
    }
    .stage-progress__percent {
        white-space: nowrap;
        font-variant-numeric: tabular-nums;
        color: #0f766e;
        font-size: 1.55rem;
        line-height: 1;
        font-weight: 800;
        min-width: 72px;
        text-align: right;
    }
    .loader {
        position: relative;
        background-color: #535353;
        border-radius: 1em;
        height: 1em;
        width: 100%;
    }
    .bar {
        position: relative;
        background-color: rgb(0, 205, 0);
        width: var(--progress-width, 0%);
        height: 100%;
        border-radius: 1em;
        transition: width 280ms ease-out;
    }
    .check-bar-container {
        position: absolute;
        left: 0;
        top: -4px;
        z-index: 2;
        display: flex;
        width: 100%;
        justify-content: space-between;
        height: 0.5em;
    }
    .check {
        border-radius: 1em;
        height: 1.5em;
        width: 1.5em;
        padding: 3px;
        background-color: #535353;
        transform: scale(0.75);
        transition: transform 180ms ease-out, background-color 180ms ease-out;
        box-sizing: border-box;
    }
    .check.is-complete {
        transform: scale(1);
        background-color: rgb(0, 205, 0);
    }
    .check.is-current {
        transform: scale(1);
        background-color: #22c55e;
        box-shadow: 0 0 0 5px rgba(34, 197, 94, 0.16);
    }
    .stage-progress__labels {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 8px;
        width: 100%;
        margin-top: 12px;
        color: #64748b;
        font-size: 0.76rem;
    }
    .stage-progress__labels span:nth-child(2),
    .stage-progress__labels span:nth-child(3) {
        text-align: center;
    }
    .stage-progress__labels span:last-child {
        text-align: right;
    }
    .stage-progress__labels .is-active {
        color: #0f766e;
        font-weight: 700;
    }
    .stage-progress.is-idle .loader {
        background-color: #bfdbfe;
    }
    .stage-progress.is-idle .bar {
        background-color: #93c5fd;
    }
    .stage-progress.is-idle .check {
        background-color: #bfdbfe;
    }
    .stage-progress.is-idle .stage-progress__percent,
    .stage-progress.is-idle .stage-progress__labels .is-active {
        color: #2563eb;
    }
    .stage-progress.is-failed .loader {
        background-color: #fecaca;
    }
    .stage-progress.is-failed .bar,
    .stage-progress.is-failed .check.is-complete,
    .stage-progress.is-failed .check.is-current {
        background-color: #ef4444;
    }
    .stage-progress.is-failed .stage-progress__percent,
    .stage-progress.is-failed .stage-progress__labels .is-active {
        color: #dc2626;
    }
    .confetti-burst {
        pointer-events: none;
        position: fixed;
        inset: 0;
        z-index: 9999;
        overflow: hidden;
    }
    .confetti-piece {
        position: absolute;
        left: 50%;
        top: 42%;
        width: 8px;
        height: 14px;
        border-radius: 2px;
        background: var(--confetti-color);
        opacity: 0;
        transform: translate(-50%, -50%) rotate(0deg);
        animation: confettiBurst 1.9s cubic-bezier(0.18, 0.86, 0.32, 1) forwards;
        animation-delay: var(--confetti-delay);
    }
    @keyframes confettiBurst {
        0% {
            opacity: 0;
            transform: translate(-50%, -50%) scale(0.45) rotate(0deg);
        }
        8% {
            opacity: 1;
        }
        38% {
            opacity: 1;
            transform: translate(calc(-50% + var(--confetti-x)), calc(-50% + var(--confetti-y))) scale(1) rotate(var(--confetti-r1));
        }
        100% {
            opacity: 0;
            transform: translate(calc(-50% + var(--confetti-x2)), calc(-50% + var(--confetti-fall))) scale(0.95) rotate(var(--confetti-r2));
        }
    }
    .fault-burst {
        pointer-events: none;
        position: fixed;
        inset: 0;
        z-index: 9999;
        display: flex;
        align-items: center;
        justify-content: center;
        animation: faultFade 1.55s ease-out forwards;
    }
    .fault-machine {
        position: relative;
        width: 158px;
        height: 108px;
        border-radius: 12px;
        background: #1f2937;
        border: 4px solid #ef4444;
        box-shadow: 0 18px 50px rgba(127, 29, 29, 0.34);
        animation: faultShake 0.12s linear 8;
    }
    .fault-machine:before {
        content: "";
        position: absolute;
        left: 24px;
        right: 24px;
        top: 24px;
        height: 16px;
        border-radius: 999px;
        background: repeating-linear-gradient(90deg, #ef4444 0 14px, transparent 14px 22px);
    }
    .fault-machine:after {
        content: "";
        position: absolute;
        left: 30px;
        right: 30px;
        bottom: 24px;
        height: 20px;
        border-radius: 4px;
        background: repeating-linear-gradient(90deg, #fca5a5 0 10px, transparent 10px 18px);
    }
    .fault-spark {
        position: absolute;
        width: 10px;
        height: 10px;
        border-radius: 2px;
        background: #f59e0b;
        opacity: 0;
        animation: faultSpark 1.1s ease-out forwards;
    }
    .fault-spark:nth-child(1) { left: -22px; top: 10px; --sx: -70px; --sy: -46px; animation-delay: 0.04s; }
    .fault-spark:nth-child(2) { right: -18px; top: 24px; --sx: 82px; --sy: -32px; animation-delay: 0.10s; }
    .fault-spark:nth-child(3) { left: 22px; top: -18px; --sx: -34px; --sy: -82px; animation-delay: 0.16s; }
    .fault-spark:nth-child(4) { right: 28px; bottom: -18px; --sx: 42px; --sy: 72px; animation-delay: 0.22s; }
    @keyframes faultShake {
        0%, 100% { transform: translateX(0) rotate(0deg); }
        25% { transform: translateX(-5px) rotate(-1.5deg); }
        75% { transform: translateX(5px) rotate(1.5deg); }
    }
    @keyframes faultSpark {
        0% { opacity: 1; transform: translate(0, 0) scale(0.8) rotate(0deg); }
        100% { opacity: 0; transform: translate(var(--sx), var(--sy)) scale(0.2) rotate(220deg); }
    }
    @keyframes faultFade {
        0%, 72% { opacity: 1; }
        100% { opacity: 0; }
    }
    .typewriter {
        --blue: #5C86FF;
        --blue-dark: #275EFE;
        --key: #fff;
        --paper: #EEF0FD;
        --text: #D3D4EC;
        --tool: #FBC56C;
        --duration: 3s;
        position: relative;
        width: 128px;
        height: 64px;
        flex: 0 0 128px;
        animation: bounce05 var(--duration) linear infinite;
    }
    .typewriter .slide {
        width: 92px;
        height: 20px;
        border-radius: 3px;
        margin-left: 14px;
        transform: translateX(14px);
        background: linear-gradient(var(--blue), var(--blue-dark));
        animation: slide05 var(--duration) ease infinite;
    }
    .typewriter .slide:before,
    .typewriter .slide:after,
    .typewriter .slide i:before {
        content: "";
        position: absolute;
        background: var(--tool);
    }
    .typewriter .slide:before {
        width: 2px;
        height: 8px;
        top: 6px;
        left: 100%;
    }
    .typewriter .slide:after {
        left: 94px;
        top: 3px;
        height: 14px;
        width: 6px;
        border-radius: 3px;
    }
    .typewriter .slide i {
        display: block;
        position: absolute;
        right: 100%;
        width: 6px;
        height: 4px;
        top: 4px;
        background: var(--tool);
    }
    .typewriter .slide i:before {
        right: 100%;
        top: -2px;
        width: 4px;
        border-radius: 2px;
        height: 14px;
    }
    .typewriter .paper {
        position: absolute;
        left: 24px;
        top: -26px;
        width: 40px;
        height: 46px;
        border-radius: 5px;
        background: var(--paper);
        transform: translateY(46px);
        animation: paper05 var(--duration) linear infinite;
    }
    .typewriter .paper:before {
        content: "";
        position: absolute;
        left: 6px;
        right: 6px;
        top: 7px;
        border-radius: 2px;
        height: 4px;
        transform: scaleY(0.8);
        background: var(--text);
        box-shadow: 0 12px 0 var(--text), 0 24px 0 var(--text), 0 36px 0 var(--text);
    }
    .typewriter .keyboard {
        width: 120px;
        height: 56px;
        margin-top: -10px;
        z-index: 1;
        position: relative;
    }
    .typewriter .keyboard:before,
    .typewriter .keyboard:after {
        content: "";
        position: absolute;
    }
    .typewriter .keyboard:before {
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        border-radius: 7px;
        background: linear-gradient(135deg, var(--blue), var(--blue-dark));
        transform: perspective(10px) rotateX(2deg);
        transform-origin: 50% 100%;
    }
    .typewriter .keyboard:after {
        left: 2px;
        top: 25px;
        width: 11px;
        height: 4px;
        border-radius: 2px;
        box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key);
        animation: keyboard05 var(--duration) linear infinite;
    }
    .stage-progress.is-static .typewriter,
    .stage-progress.is-static .typewriter * {
        animation: none !important;
    }
    .stage-progress.is-static .typewriter .slide:before,
    .stage-progress.is-static .typewriter .slide:after,
    .stage-progress.is-static .typewriter .slide i:before,
    .stage-progress.is-static .typewriter .paper:before,
    .stage-progress.is-static .typewriter .keyboard:before,
    .stage-progress.is-static .typewriter .keyboard:after {
        animation: none !important;
    }
    .stage-progress.is-static .typewriter {
        transform: translateY(0);
    }
    .stage-progress.is-static .typewriter .slide {
        transform: translateX(14px);
    }
    .stage-progress.is-static .typewriter .paper {
        transform: translateY(46px);
    }
    .stage-progress.is-static .typewriter .keyboard:after {
        animation: none !important;
        box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key);
    }
    @media (max-width: 760px) {
        .stage-progress__row {
            gap: 12px;
        }
        .typewriter {
            transform: scale(0.78);
            transform-origin: center;
            flex-basis: 104px;
        }
        .stage-progress__percent {
            font-size: 1.25rem;
            min-width: 58px;
        }
    }
    @keyframes bounce05 {
        85%, 92%, 100% { transform: translateY(0); }
        89% { transform: translateY(-4px); }
        95% { transform: translateY(2px); }
    }
    @keyframes slide05 {
        5% { transform: translateX(14px); }
        15%, 30% { transform: translateX(6px); }
        40%, 55% { transform: translateX(0); }
        65%, 70% { transform: translateX(-4px); }
        80%, 89% { transform: translateX(-12px); }
        100% { transform: translateX(14px); }
    }
    @keyframes paper05 {
        5% { transform: translateY(46px); }
        20%, 30% { transform: translateY(34px); }
        40%, 55% { transform: translateY(22px); }
        65%, 70% { transform: translateY(10px); }
        80%, 85% { transform: translateY(0); }
        92%, 100% { transform: translateY(46px); }
    }
    @keyframes keyboard05 {
        5%, 12%, 21%, 30%, 39%, 48%, 57%, 66%, 75%, 84% {
            box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key);
        }
        9% { box-shadow: 15px 2px 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        18% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 2px 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        27% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 12px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        36% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 12px 0 var(--key), 60px 12px 0 var(--key), 68px 12px 0 var(--key), 83px 10px 0 var(--key); }
        45% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 2px 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        54% { box-shadow: 15px 0 0 var(--key), 30px 2px 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        63% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 12px 0 var(--key); }
        72% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 2px 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 10px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
        81% { box-shadow: 15px 0 0 var(--key), 30px 0 0 var(--key), 45px 0 0 var(--key), 60px 0 0 var(--key), 75px 0 0 var(--key), 90px 0 0 var(--key), 22px 10px 0 var(--key), 37px 12px 0 var(--key), 52px 10px 0 var(--key), 60px 10px 0 var(--key), 68px 10px 0 var(--key), 83px 10px 0 var(--key); }
    }
</style>
        """,
        unsafe_allow_html=True,
    )


def _notify_success(message: str) -> None:
    toast_fn = getattr(st, "toast", None)
    if callable(toast_fn):
        toast_fn(message, icon="✅")
    else:
        st.success(message)


def _notify_error(message: str) -> None:
    toast_fn = getattr(st, "toast", None)
    if callable(toast_fn):
        toast_fn(message, icon="❌")
    else:
        st.error(message)


def _default_api_provider_config() -> dict[str, str]:
    return {
        "provider_name": "OpenAI",
        "base_url": os.getenv("OPENAI_BASE_URL", API_PROVIDER_PRESETS["OpenAI"]["base_url"]),
        "api_key": os.getenv("OPENAI_API_KEY", ""),
        "model": os.getenv("AUTOPDFTRANSLATOR_MODEL", API_PROVIDER_PRESETS["OpenAI"]["model"]),
        "prompt_hint": os.getenv("AUTOPDFTRANSLATOR_PROMPT_HINT", ""),
        "provider_impl": "openai-compatible",
    }


def _default_local_provider_config() -> dict[str, str]:
    return {
        "provider_name": "Ollama",
        "base_url": LOCAL_MODEL_PRESETS["Ollama"]["base_url"],
        "api_key": "",
        "model": LOCAL_MODEL_PRESETS["Ollama"]["model"],
        "prompt_hint": "",
        "provider_impl": "openai-compatible",
    }


def _default_export_output_dir() -> str:
    return str((APP_STORAGE_ROOT / "outputs").resolve())


def _is_streamlit_cloud_runtime() -> bool:
    markers = [
        "STREAMLIT_SHARING_MODE",
        "STREAMLIT_CLOUD",
        "STREAMLIT_RUNTIME",
    ]
    if any(os.getenv(marker) for marker in markers):
        return True
    resolved = str(Path(__file__).resolve()).replace("\\", "/").lower()
    cwd = str(Path.cwd().resolve()).replace("\\", "/").lower()
    return resolved.startswith("/mount/src/") or cwd.startswith("/mount/src/")


def _server_path_saving_enabled() -> bool:
    return not _is_streamlit_cloud_runtime()


def _snap_to_adobe_font_size(value: Any) -> float:
    try:
        numeric = float(value)
    except Exception:
        numeric = DEFAULT_ANNOTATION_FONT_SIZE
    return min(ADOBE_ANNOTATION_FONT_SIZES, key=lambda x: abs(x - numeric))


def _load_ui_settings() -> dict[str, Any]:
    if not UI_CONFIG_PATH.exists():
        return {}
    try:
        data = json.loads(UI_CONFIG_PATH.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return data
    except Exception:
        return {}
    return {}


def _save_ui_settings(settings: dict[str, Any]) -> None:
    UI_CONFIG_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_translation_history() -> list[dict[str, str]]:
    if not TRANSLATION_HISTORY_PATH.exists():
        return []
    try:
        data = json.loads(TRANSLATION_HISTORY_PATH.read_text(encoding="utf-8"))
        if isinstance(data, list):
            out: list[dict[str, str]] = []
            for item in data:
                if isinstance(item, dict):
                    out.append(
                        {
                            "id": str(item.get("id", "")).strip() or str(uuid.uuid4()),
                            "translated_at": str(item.get("translated_at", "")),
                            "source_file": str(item.get("source_file", "")),
                            "source_path": str(item.get("source_path", "")),
                            "translated_file": str(item.get("translated_file", "")),
                            "translated_path": str(item.get("translated_path", "")),
                            "mode": str(item.get("mode", "")),
                            "result": str(item.get("result", "")),
                            "task_duration": str(item.get("task_duration", "")),
                        }
                    )
            return out
    except Exception:
        return []
    return []


def _save_translation_history(history: list[dict[str, str]]) -> None:
    TRANSLATION_HISTORY_PATH.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_translation_corpus() -> list[dict[str, str]]:
    if not TRANSLATION_CORPUS_PATH.exists():
        return []
    try:
        data = json.loads(TRANSLATION_CORPUS_PATH.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            return []
        entries: list[dict[str, str]] = []
        for item in data:
            if not isinstance(item, dict):
                continue
            text = str(item.get("text", "")).strip()
            filename = str(item.get("filename", "")).strip()
            if not text or not filename:
                continue
            entries.append(
                {
                    "id": str(item.get("id", "")).strip() or str(uuid.uuid4()),
                    "filename": filename,
                    "created_at": str(item.get("created_at", "")),
                    "text": text[:MAX_CORPUS_ENTRY_CHARS],
                }
            )
        return entries
    except Exception:
        return []


def _save_translation_corpus(entries: list[dict[str, str]]) -> None:
    TRANSLATION_CORPUS_PATH.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")


def _empty_translation_memory() -> dict[str, Any]:
    return {
        "terms": [],
        "style_rules": [],
        "examples": [],
        "updated_at": "",
    }


def _load_translation_memory() -> dict[str, Any]:
    if not TRANSLATION_MEMORY_PATH.exists():
        return _empty_translation_memory()
    try:
        data = json.loads(TRANSLATION_MEMORY_PATH.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return _empty_translation_memory()
        memory = _empty_translation_memory()
        terms = data.get("terms", [])
        if isinstance(terms, list):
            memory["terms"] = [
                {
                    "source": str(item.get("source", "")).strip(),
                    "target": str(item.get("target", "")).strip(),
                    "full_form": str(item.get("full_form", "")).strip(),
                    "source_file": str(item.get("source_file", "")).strip(),
                }
                for item in terms
                if isinstance(item, dict)
                and str(item.get("source", "")).strip()
            ]
        for key in ("style_rules", "examples"):
            values = data.get(key, [])
            if isinstance(values, list):
                memory[key] = [str(value).strip() for value in values if str(value).strip()]
        memory["updated_at"] = str(data.get("updated_at", ""))
        return memory
    except Exception:
        return _empty_translation_memory()


def _save_translation_memory(memory: dict[str, Any]) -> None:
    TRANSLATION_MEMORY_PATH.write_text(json.dumps(memory, ensure_ascii=False, indent=2), encoding="utf-8")


def _decode_reference_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _clean_reference_text(text: str) -> str:
    lines: list[str] = []
    for line in re.split(r"[\r\n]+", text):
        line = re.sub(r"\s+", " ", line).strip()
        if line:
            lines.append(line)
    return "\n".join(lines).strip()


def _extract_text_from_xml_bytes(data: bytes) -> str:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return ""
    parts: list[str] = []
    for node in root.iter():
        if node.text and node.text.strip():
            parts.append(node.text.strip())
    return "\n".join(parts)


def _extract_text_from_office_zip(data: bytes, suffix: str) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        names = archive.namelist()
        if suffix == ".docx":
            selected = [name for name in names if name == "word/document.xml"]
            selected += [name for name in names if name.startswith("word/") and name.endswith(("header1.xml", "footer1.xml"))]
        elif suffix == ".pptx":
            selected = sorted(name for name in names if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        elif suffix == ".xlsx":
            selected = [name for name in names if name == "xl/sharedStrings.xml"]
            selected += sorted(name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"))
        else:
            selected = []

        chunks: list[str] = []
        for name in selected:
            try:
                text = _extract_text_from_xml_bytes(archive.read(name))
            except Exception:
                text = ""
            if text.strip():
                chunks.append(text)
    return "\n".join(chunks)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _row_to_term_line(values: list[str]) -> str:
    for idx in range(len(values) - 1):
        source = values[idx].strip()
        target = values[idx + 1].strip()
        if _looks_like_source_term(source) and _looks_like_target_term(target):
            return f"{source} => {target}"
    return ""


def _extract_text_from_xlsx_bytes(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return _extract_text_from_office_zip(data, ".xlsx")

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [_cell_to_text(value) for value in row]
            values = [value for value in values if value]
            if not values:
                continue
            term_line = _row_to_term_line(values)
            if term_line:
                lines.append(term_line)
            else:
                lines.extend(values)
    return "\n".join(lines)


def _find_header_index(headers: list[str], candidates: set[str]) -> int | None:
    for idx, header in enumerate(headers):
        if header.strip() in candidates:
            return idx
    return None


def _extract_memory_from_xlsx_bytes(filename: str, data: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return {"terms": [], "examples": []}

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    terms: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        header_row: list[str] = []
        english_idx: int | None = None
        chinese_idx: int | None = None
        abbr_idx: int | None = None
        full_idx: int | None = None

        for row in sheet.iter_rows(values_only=True):
            values = [_cell_to_text(value) for value in row]
            if not any(values):
                continue

            if not header_row:
                possible_english_idx = _find_header_index(values, {"英文词汇", "英文", "English", "Term"})
                possible_chinese_idx = _find_header_index(values, {"中文词汇", "中文", "Chinese", "Translation"})
                possible_abbr_idx = _find_header_index(values, {"缩写", "Abbreviation", "Abbr"})
                possible_full_idx = _find_header_index(values, {"完整", "完整拼写", "Full", "Full form"})
                if possible_english_idx is not None and possible_chinese_idx is not None:
                    header_row = values
                    english_idx = possible_english_idx
                    chinese_idx = possible_chinese_idx
                    continue
                if possible_abbr_idx is not None and possible_full_idx is not None:
                    header_row = values
                    abbr_idx = possible_abbr_idx
                    full_idx = possible_full_idx
                    continue

            if english_idx is not None and chinese_idx is not None:
                source = values[english_idx] if english_idx < len(values) else ""
                target = values[chinese_idx] if chinese_idx < len(values) else ""
                if _looks_like_source_term(source) and _looks_like_target_term(target):
                    terms.append(
                        {
                            "source": source,
                            "target": target,
                            "full_form": "",
                            "source_file": filename,
                        }
                    )
            elif abbr_idx is not None and full_idx is not None:
                source = values[abbr_idx] if abbr_idx < len(values) else ""
                full_form = values[full_idx] if full_idx < len(values) else ""
                if _looks_like_source_term(source) and _looks_like_source_term(full_form):
                    terms.append(
                        {
                            "source": source,
                            "target": "",
                            "full_form": full_form,
                            "source_file": filename,
                        }
                    )
    return {"terms": terms, "examples": []}


def _extract_text_from_pdf_bytes(data: bytes) -> str:
    import fitz

    chunks: list[str] = []
    with fitz.open(stream=data, filetype="pdf") as doc:
        for page in doc:
            text = page.get_text("text").strip()
            if text:
                chunks.append(text)
    return "\n".join(chunks)


def _extract_reference_text(filename: str, data: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        text = _extract_text_from_pdf_bytes(data)
    elif suffix == ".xlsx":
        text = _extract_text_from_xlsx_bytes(data)
    elif suffix in {".docx", ".pptx"}:
        text = _extract_text_from_office_zip(data, suffix)
    elif suffix in {".txt", ".md", ".csv"}:
        text = _decode_reference_bytes(data)
    elif suffix in {".doc", ".ppt", ".xls"}:
        raise ValueError("旧版 Office 格式暂不支持自动读取，请另存为 docx/pptx/xlsx 后再导入。")
    else:
        raise ValueError(f"暂不支持的参考文件格式: {suffix or filename}")
    return _clean_reference_text(text)[:MAX_CORPUS_ENTRY_CHARS]


def _looks_like_source_term(value: str) -> bool:
    return bool(re.search(r"[A-Za-z]", value)) and 1 <= len(value.strip()) <= 90


def _looks_like_target_term(value: str) -> bool:
    value = value.strip()
    return bool(value) and len(value) <= 120 and not re.fullmatch(r"[\d\s.,;:/\\|()\-]+", value)


def _extract_memory_from_reference(filename: str, text: str) -> dict[str, Any]:
    terms: list[dict[str, str]] = []
    examples: list[str] = []
    separators = r"(?:=>|->|=|\t|：|:)"
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if len(examples) < 4 and len(line) <= MAX_MEMORY_EXAMPLE_CHARS:
            examples.append(line)
        match = re.match(rf"^(.{{1,90}}?)\s*{separators}\s*(.{{1,120}})$", line)
        if not match and line.count(",") == 1:
            match = re.match(r"^(.{1,90}?),\s*(.{1,120})$", line)
        if not match:
            continue
        source = match.group(1).strip(" -•\t")
        target = match.group(2).strip(" -•\t")
        if _looks_like_source_term(source) and _looks_like_target_term(target):
            terms.append(
                {
                    "source": source,
                    "target": target,
                    "full_form": "",
                    "source_file": filename,
                }
            )
    return {"terms": terms, "examples": examples}


def _merge_translation_memory(
    current: dict[str, Any],
    addition: dict[str, Any],
) -> dict[str, Any]:
    memory = {
        "terms": list(current.get("terms", [])) if isinstance(current.get("terms", []), list) else [],
        "style_rules": list(current.get("style_rules", [])) if isinstance(current.get("style_rules", []), list) else [],
        "examples": list(current.get("examples", [])) if isinstance(current.get("examples", []), list) else [],
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    term_map: dict[str, dict[str, str]] = {}
    for item in memory["terms"]:
        if not isinstance(item, dict):
            continue
        source = str(item.get("source", "")).strip()
        target = str(item.get("target", "")).strip()
        full_form = str(item.get("full_form", "")).strip()
        if source:
            term_map[source.lower()] = {
                "source": source,
                "target": target,
                "full_form": full_form,
                "source_file": str(item.get("source_file", "")).strip(),
            }
    for item in addition.get("terms", []):
        if not isinstance(item, dict):
            continue
        source = str(item.get("source", "")).strip()
        target = str(item.get("target", "")).strip()
        full_form = str(item.get("full_form", "")).strip()
        if source:
            term_map[source.lower()] = {
                "source": source,
                "target": target,
                "full_form": full_form,
                "source_file": str(item.get("source_file", "")).strip(),
            }
    memory["terms"] = sorted(term_map.values(), key=lambda item: item["source"].lower())

    seen_examples = {str(value).strip() for value in memory["examples"]}
    for example in addition.get("examples", []):
        example = str(example).strip()
        if example and example not in seen_examples:
            memory["examples"].append(example[:MAX_MEMORY_EXAMPLE_CHARS])
            seen_examples.add(example)
    memory["examples"] = memory["examples"][-20:]
    return memory


def _rebuild_memory_from_corpus(entries: list[dict[str, str]]) -> dict[str, Any]:
    memory = _empty_translation_memory()
    for entry in entries:
        addition = _extract_memory_from_reference(
            str(entry.get("filename", "reference")),
            str(entry.get("text", "")),
        )
        memory = _merge_translation_memory(memory, addition)
    return memory


def _extract_memory_from_uploaded_reference(filename: str, data: bytes, text: str) -> dict[str, Any]:
    if Path(filename).suffix.lower() == ".xlsx":
        structured = _extract_memory_from_xlsx_bytes(filename, data)
        if structured.get("terms"):
            return structured
    return _extract_memory_from_reference(filename, text)


def _build_memory_env_payload(memory: dict[str, Any]) -> str:
    payload = {
        "terms": memory.get("terms", []),
        "style_rules": memory.get("style_rules", []),
        "examples": memory.get("examples", []),
        "updated_at": memory.get("updated_at", ""),
    }
    return json.dumps(payload, ensure_ascii=False)


def _clear_runtime_memory_env() -> None:
    os.environ.pop("AUTOPDFTRANSLATOR_MEMORY_JSON", None)
    os.environ.pop("AUTOPDFTRANSLATOR_MEMORY_JSON_PATH", None)


def _apply_runtime_memory(memory: dict[str, Any]) -> None:
    payload = _build_memory_env_payload(memory)
    if len(payload) <= 30000:
        os.environ["AUTOPDFTRANSLATOR_MEMORY_JSON"] = payload
        os.environ.pop("AUTOPDFTRANSLATOR_MEMORY_JSON_PATH", None)
        return

    RUNTIME_MEMORY_PATH.write_text(payload, encoding="utf-8")
    os.environ.pop("AUTOPDFTRANSLATOR_MEMORY_JSON", None)
    os.environ["AUTOPDFTRANSLATOR_MEMORY_JSON_PATH"] = str(RUNTIME_MEMORY_PATH)


def _post_chat_for_term_completion(
    config: dict[str, str],
    messages: list[dict[str, str]],
    *,
    max_tokens: int,
) -> str:
    if requests is None:
        raise RuntimeError("requests is required for AI term completion")
    base_url = config.get("base_url", "").strip().rstrip("/")
    model = config.get("model", "").strip()
    api_key = _normalize_bearer_api_key(config.get("api_key", ""))
    if not base_url or not model:
        raise RuntimeError("Base URL and Model are required for AI term completion.")
    headers = {"Content-Type": "application/json"}
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    base_lower = base_url.lower()
    model_lower = model.lower()
    temperature = 1.0 if "moonshot.cn" in base_lower or model_lower.startswith("kimi-k2.5") else 0.0
    payload = {
        "model": model,
        "messages": messages,
        "temperature": temperature,
        "max_tokens": max_tokens,
    }
    if "minimax.io" in base_lower or "minimaxi.com" in base_lower:
        payload["reasoning_split"] = True
    response = requests.post(
        f"{base_url}/chat/completions",
        headers=headers,
        json=payload,
        timeout=300,
    )
    if response.status_code >= 400:
        raise RuntimeError(f"HTTP {response.status_code}: {response.text[:800]}")
    data = response.json()
    choices = data.get("choices", [])
    if choices and isinstance(choices[0], dict):
        message = choices[0].get("message", {})
        if isinstance(message, dict):
            content = message.get("content", "")
            if isinstance(content, str) and content.strip():
                return re.sub(r"<think>.*?</think>", "", content, flags=re.IGNORECASE | re.DOTALL).strip()
            reasoning = message.get("reasoning_content", "")
            if isinstance(reasoning, str) and reasoning.strip():
                raise RuntimeError(
                    "Model returned reasoning_content but no final JSON content. "
                    "The output token budget may be too small for this model; try again or use a non-reasoning model. "
                    f"reasoning_preview={reasoning[:500]}"
                )
    raise RuntimeError(f"Model returned no text content: {str(data)[:800]}")


def _parse_term_completion_json(raw: str) -> list[dict[str, str]]:
    raw = raw.strip()
    if raw.startswith("```"):
        raw = re.sub(r"^```(?:json)?\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        partial = _parse_partial_term_completion_json(raw)
        if partial:
            return partial
        raise exc
    if not isinstance(data, list):
        return []
    out: list[dict[str, str]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        source = str(item.get("source", "")).strip()
        target = str(item.get("target", "")).strip()
        if source and target:
            out.append({"source": source, "target": target})
    return out


def _parse_partial_term_completion_json(raw: str) -> list[dict[str, str]]:
    decoder = json.JSONDecoder()
    out: list[dict[str, str]] = []
    idx = 0
    while idx < len(raw):
        object_start = raw.find("{", idx)
        if object_start < 0:
            break
        try:
            parsed, end = decoder.raw_decode(raw[object_start:])
        except json.JSONDecodeError:
            idx = object_start + 1
            continue
        if isinstance(parsed, dict):
            source = str(parsed.get("source", "")).strip()
            target = str(parsed.get("target", "")).strip()
            if source and target:
                out.append({"source": source, "target": target})
        idx = object_start + max(end, 1)
    if out:
        return out

    # Last-resort recovery for model output that is nearly JSON but has a broken later string.
    pattern = re.compile(
        r'\{\s*"source"\s*:\s*"((?:\\.|[^"\\])*)"\s*,\s*"target"\s*:\s*"((?:\\.|[^"\\])*)"\s*\}',
        re.DOTALL,
    )
    for match in pattern.finditer(raw):
        try:
            source = json.loads(f'"{match.group(1)}"').strip()
            target = json.loads(f'"{match.group(2)}"').strip()
        except Exception:
            continue
        if source and target:
            out.append({"source": source, "target": target})
    return out


def _build_term_completion_context(memory: dict[str, Any]) -> str:
    terms = memory.get("terms", [])
    if not isinstance(terms, list):
        return ""
    known = [
        item
        for item in terms
        if isinstance(item, dict)
        and str(item.get("source", "")).strip()
        and str(item.get("target", "")).strip()
    ][:80]
    if not known:
        return ""
    lines = [
        f"- {item.get('source', '')} => {item.get('target', '')}"
        for item in known
    ]
    return "\n".join(lines)


def _complete_missing_memory_targets(memory: dict[str, Any], config: dict[str, str]) -> int:
    terms = memory.get("terms", [])
    if not isinstance(terms, list):
        return 0
    pending_indices = [
        idx
        for idx, item in enumerate(terms)
        if isinstance(item, dict)
        and str(item.get("source", "")).strip()
        and not str(item.get("target", "")).strip()
    ]
    if not pending_indices:
        return 0

    completed = 0
    known_context = _build_term_completion_context(memory)
    payload = [
        {
            "source": str(terms[idx].get("source", "")).strip(),
            "full_form": str(terms[idx].get("full_form", "")).strip(),
        }
        for idx in pending_indices
    ]
    prompt = (
        "You are building a professional bilingual glossary for architecture, construction drawings, "
        "and architectural visualization review comments.\n"
        "First infer the domain from the known glossary context, then complete ALL missing Chinese terms in one pass.\n"
        "Do NOT machine-translate literally. Use established professional terminology in Chinese.\n"
        "For abbreviations, translate the full form by its architecture/engineering meaning, not by word-by-word literal meaning. "
        "Example: 'Above finished floor' should be '地板标高以上', not '完成地面以上'.\n"
        "Return strict JSON array only. Each item must be {\"source\":\"...\",\"target\":\"...\"}. "
        "Keep source exactly the same as input source. Return one output item for every input item. No extra prose.\n\n"
        f"Known glossary context:\n{known_context or '(none)'}\n\n"
        f"Terms to complete:\n{json.dumps(payload, ensure_ascii=False, separators=(',', ':'))}"
    )
    max_tokens = max(6000, min(20000, 2200 + 90 * len(payload)))
    raw = _post_chat_for_term_completion(
        config,
        [
            {"role": "system", "content": "You are a senior architecture and construction terminology translator. Return final JSON only."},
            {"role": "user", "content": prompt},
        ],
        max_tokens=max_tokens,
    )
    completed_items = _parse_term_completion_json(raw)
    completed_map = {item["source"]: item["target"] for item in completed_items}
    for idx in pending_indices:
        source = str(terms[idx].get("source", "")).strip()
        translated = completed_map.get(source, "").strip()
        if translated:
            terms[idx]["target"] = translated
            completed += 1
    memory["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return completed


def _build_effective_prompt_hint(base_prompt: str) -> str:
    return base_prompt.strip()


def _persist_ui_settings() -> None:
    payload = {
        "runtime_provider_group": st.session_state.get("runtime_provider_group", "API Providers"),
        "api_provider_config": st.session_state.get("api_provider_config", _default_api_provider_config()),
        "local_provider_config": st.session_state.get("local_provider_config", _default_local_provider_config()),
        "usd_to_cny_rate": float(st.session_state.get("usd_to_cny_rate", DEFAULT_USD_TO_CNY_RATE)),
        "export_output_dir": str(st.session_state.get("export_output_dir", _default_export_output_dir())),
        "typed_text_threshold": int(st.session_state.get("typed_text_threshold", 120)),
        "dpi_for_vision": int(st.session_state.get("dpi_for_vision", 220)),
        "font_size": float(st.session_state.get("font_size", DEFAULT_ANNOTATION_FONT_SIZE)),
        "placement_label": str(st.session_state.get("placement_label", "top-left stacked")),
        "max_pages_raw": str(st.session_state.get("max_pages_raw", "")),
        "prompt_overhead_tokens": int(st.session_state.get("prompt_overhead_tokens", 85)),
        "vision_input_tokens_per_page": int(st.session_state.get("vision_input_tokens_per_page", 1200)),
        "vision_output_tokens_per_page": int(st.session_state.get("vision_output_tokens_per_page", 220)),
        "input_cost_per_1m": float(st.session_state.get("input_cost_per_1m", 0.40)),
        "output_cost_per_1m": float(st.session_state.get("output_cost_per_1m", 1.60)),
        "fast_mode_enabled": bool(st.session_state.get("fast_mode_enabled", False)),
        "fast_mode_dpi": int(st.session_state.get("fast_mode_dpi", 150)),
        "fast_mode_timeout_seconds": int(st.session_state.get("fast_mode_timeout_seconds", 180)),
        "fast_mode_max_tokens": int(st.session_state.get("fast_mode_max_tokens", 2048)),
        "fast_mode_retry_attempts": int(st.session_state.get("fast_mode_retry_attempts", 2)),
        "odl_fallback_enabled": bool(st.session_state.get("odl_fallback_enabled", False)),
        "corpus_enabled": bool(st.session_state.get("corpus_enabled", True)),
    }
    try:
        _save_ui_settings(payload)
    except Exception as exc:
        st.warning(f"Failed to persist UI settings: {exc}")


def _init_state() -> None:
    saved = _load_ui_settings()
    saved_history = _load_translation_history()
    saved_corpus = _load_translation_corpus()
    saved_memory = _load_translation_memory()
    if saved_corpus and not saved_memory.get("terms") and not saved_memory.get("examples"):
        saved_memory = _rebuild_memory_from_corpus(saved_corpus)
        try:
            _save_translation_memory(saved_memory)
        except Exception:
            pass
    api_config = saved.get("api_provider_config", _default_api_provider_config())
    local_config = saved.get("local_provider_config", _default_local_provider_config())
    provider_group = saved.get("runtime_provider_group", "API Providers")
    usd_to_cny_rate = saved.get("usd_to_cny_rate", DEFAULT_USD_TO_CNY_RATE)
    export_output_dir = saved.get("export_output_dir", _default_export_output_dir())
    typed_text_threshold = saved.get("typed_text_threshold", 120)
    dpi_for_vision = saved.get("dpi_for_vision", 220)
    font_size = saved.get("font_size", DEFAULT_ANNOTATION_FONT_SIZE)
    placement_label = saved.get("placement_label", "top-left stacked")
    max_pages_raw = saved.get("max_pages_raw", "")
    prompt_overhead_tokens = saved.get("prompt_overhead_tokens", 85)
    vision_input_tokens_per_page = saved.get("vision_input_tokens_per_page", 1200)
    vision_output_tokens_per_page = saved.get("vision_output_tokens_per_page", 220)
    input_cost_per_1m = saved.get("input_cost_per_1m", 0.40)
    output_cost_per_1m = saved.get("output_cost_per_1m", 1.60)
    fast_mode_enabled = bool(saved.get("fast_mode_enabled", False))
    fast_mode_dpi = saved.get("fast_mode_dpi", 150)
    fast_mode_timeout_seconds = saved.get("fast_mode_timeout_seconds", 180)
    fast_mode_max_tokens = saved.get("fast_mode_max_tokens", 2048)
    fast_mode_retry_attempts = saved.get("fast_mode_retry_attempts", 2)
    odl_fallback_enabled = bool(saved.get("odl_fallback_enabled", False))
    corpus_enabled = bool(saved.get("corpus_enabled", True))

    if provider_group not in {"API Providers", "Local Models"}:
        provider_group = "API Providers"
    if not isinstance(api_config, dict):
        api_config = _default_api_provider_config()
    else:
        api_config = {**_default_api_provider_config(), **api_config}
    if not isinstance(local_config, dict):
        local_config = _default_local_provider_config()
    else:
        local_config = {**_default_local_provider_config(), **local_config}
    try:
        usd_to_cny_rate = float(usd_to_cny_rate)
    except Exception:
        usd_to_cny_rate = DEFAULT_USD_TO_CNY_RATE
    if not isinstance(export_output_dir, str) or not export_output_dir.strip():
        export_output_dir = _default_export_output_dir()
    try:
        fast_mode_dpi = int(fast_mode_dpi)
    except Exception:
        fast_mode_dpi = 150
    try:
        fast_mode_timeout_seconds = int(fast_mode_timeout_seconds)
    except Exception:
        fast_mode_timeout_seconds = 180
    try:
        fast_mode_max_tokens = int(fast_mode_max_tokens)
    except Exception:
        fast_mode_max_tokens = 2048
    try:
        fast_mode_retry_attempts = int(fast_mode_retry_attempts)
    except Exception:
        fast_mode_retry_attempts = 2
    try:
        typed_text_threshold = int(typed_text_threshold)
    except Exception:
        typed_text_threshold = 120
    try:
        dpi_for_vision = int(dpi_for_vision)
    except Exception:
        dpi_for_vision = 220
    try:
        font_size = float(font_size)
    except Exception:
        font_size = DEFAULT_ANNOTATION_FONT_SIZE
    font_size = _snap_to_adobe_font_size(font_size)
    if placement_label not in PLACEMENT_OPTIONS:
        placement_label = "top-left stacked"
    if not isinstance(max_pages_raw, str):
        max_pages_raw = str(max_pages_raw or "")
    try:
        prompt_overhead_tokens = int(prompt_overhead_tokens)
    except Exception:
        prompt_overhead_tokens = 85
    try:
        vision_input_tokens_per_page = int(vision_input_tokens_per_page)
    except Exception:
        vision_input_tokens_per_page = 1200
    try:
        vision_output_tokens_per_page = int(vision_output_tokens_per_page)
    except Exception:
        vision_output_tokens_per_page = 220
    try:
        input_cost_per_1m = float(input_cost_per_1m)
    except Exception:
        input_cost_per_1m = 0.40
    try:
        output_cost_per_1m = float(output_cost_per_1m)
    except Exception:
        output_cost_per_1m = 1.60

    defaults: dict[str, Any] = {
        "run_message": "",
        "run_error": "",
        "artifacts": [],
        "batch_zip_bytes": b"",
        "batch_zip_name": "",
        "runtime_provider_group": provider_group,
        "api_provider_config": api_config,
        "local_provider_config": local_config,
        "usd_to_cny_rate": usd_to_cny_rate,
        "export_output_dir": export_output_dir,
        "typed_text_threshold": max(1, min(400, typed_text_threshold)),
        "dpi_for_vision": max(72, min(400, dpi_for_vision)),
        "font_size": _snap_to_adobe_font_size(font_size),
        "placement_label": placement_label,
        "max_pages_raw": max_pages_raw,
        "prompt_overhead_tokens": max(1, prompt_overhead_tokens),
        "vision_input_tokens_per_page": max(1, vision_input_tokens_per_page),
        "vision_output_tokens_per_page": max(1, vision_output_tokens_per_page),
        "input_cost_per_1m": max(0.0, input_cost_per_1m),
        "output_cost_per_1m": max(0.0, output_cost_per_1m),
        "task_progress_percent": 0,
        "task_progress_text": "任务未开始",
        "task_progress_stage": "idle",
        "confetti_run_id": "",
        "confetti_rendered_run_id": "",
        "failure_animation_run_id": "",
        "failure_animation_rendered_run_id": "",
        "browser_download_run_id": "",
        "browser_download_rendered_run_id": "",
        "translation_running": False,
        "translation_cancel_requested": False,
        "save_path_message": "",
        "preflight_reports": [],
        "preflight_key": "",
        "uploaded_files_signature": "",
        "history_delete_mode": False,
        "history_search_query": "",
        "history_clear_search_requested": False,
        "translation_history": saved_history,
        "translation_corpus": saved_corpus,
        "translation_memory": saved_memory,
        "corpus_enabled": corpus_enabled,
        "fast_mode_enabled": fast_mode_enabled,
        "fast_mode_dpi": max(96, min(220, fast_mode_dpi)),
        "fast_mode_timeout_seconds": max(60, min(360, fast_mode_timeout_seconds)),
        "fast_mode_max_tokens": max(512, min(8192, fast_mode_max_tokens)),
        "fast_mode_retry_attempts": max(1, min(3, fast_mode_retry_attempts)),
        "odl_fallback_enabled": odl_fallback_enabled,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

    if "api_provider_name_ui" not in st.session_state:
        st.session_state["api_provider_name_ui"] = st.session_state["api_provider_config"]["provider_name"]
    if "api_base_url_ui" not in st.session_state:
        st.session_state["api_base_url_ui"] = st.session_state["api_provider_config"]["base_url"]
    if "api_api_key_ui" not in st.session_state:
        st.session_state["api_api_key_ui"] = st.session_state["api_provider_config"]["api_key"]
    if "api_model_ui" not in st.session_state:
        st.session_state["api_model_ui"] = st.session_state["api_provider_config"]["model"]
    if "api_prompt_ui" not in st.session_state:
        st.session_state["api_prompt_ui"] = st.session_state["api_provider_config"]["prompt_hint"]

    if "local_provider_name_ui" not in st.session_state:
        st.session_state["local_provider_name_ui"] = st.session_state["local_provider_config"]["provider_name"]
    if "local_base_url_ui" not in st.session_state:
        st.session_state["local_base_url_ui"] = st.session_state["local_provider_config"]["base_url"]
    if "local_api_key_ui" not in st.session_state:
        st.session_state["local_api_key_ui"] = st.session_state["local_provider_config"]["api_key"]
    if "local_model_ui" not in st.session_state:
        st.session_state["local_model_ui"] = st.session_state["local_provider_config"]["model"]
    if "local_prompt_ui" not in st.session_state:
        st.session_state["local_prompt_ui"] = st.session_state["local_provider_config"]["prompt_hint"]


def _on_api_provider_name_change() -> None:
    name = str(st.session_state.get("api_provider_name_ui", "OpenAI"))
    preset = API_PROVIDER_PRESETS.get(name, API_PROVIDER_PRESETS["OpenAI"])
    st.session_state["api_base_url_ui"] = preset["base_url"]
    st.session_state["api_model_ui"] = preset["model"]


def _on_local_provider_name_change() -> None:
    name = str(st.session_state.get("local_provider_name_ui", "Ollama"))
    preset = LOCAL_MODEL_PRESETS.get(name, LOCAL_MODEL_PRESETS["Ollama"])
    st.session_state["local_base_url_ui"] = preset["base_url"]
    st.session_state["local_model_ui"] = preset["model"]


def _results_to_json(results: list[Any]) -> bytes:
    payload: list[str] = []
    for result in results:
        for item in result.comments:
            translated = (item.text_translated or "").strip()
            if translated:
                payload.append(translated)
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")


def _build_page_stats(results: list[Any]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for page in results:
        execution_path = "Vision" if page.used_vision else "Text"
        translation_errors = sum(1 for item in page.comments if item.metadata.get("translation_error"))
        rows.append(
            {
                "Page": page.page_index + 1,
                "Comments": len(page.comments),
                "Translation errors": translation_errors,
                "Text chars": page.text_chars,
                "Vision used": page.used_vision,
                "Execution": execution_path,
                "Notes": "; ".join(page.notes) if page.notes else "",
            }
        )
    return rows


def _build_zip(artifacts: list[dict[str, Any]]) -> tuple[bytes, str]:
    if not artifacts:
        return b"", ""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for artifact in artifacts:
            zf.writestr(artifact["output_pdf_name"], artifact["output_pdf_bytes"])
            zf.writestr(
                artifact["output_pdf_name"].replace(".pdf", ".results.json"),
                artifact["results_json_bytes"],
            )
    return zip_buffer.getvalue(), "translated_pdfs_bundle.zip"


def _save_artifacts_to_path(
    artifacts: list[dict[str, Any]],
    target_dir_raw: str,
    batch_zip_bytes: bytes,
    batch_zip_name: str,
    *,
    save_json: bool = True,
    save_zip: bool = True,
) -> tuple[bool, str]:
    if not artifacts:
        return False, "没有可保存的翻译结果。"
    target_dir = Path(target_dir_raw).expanduser()
    if not target_dir_raw.strip():
        return False, "请先填写保存目录。"
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        saved_count = 0
        saved_pdf = 0
        saved_json = 0
        saved_zip = 0
        for artifact in artifacts:
            pdf_path = target_dir / artifact["output_pdf_name"]
            pdf_path.write_bytes(artifact["output_pdf_bytes"])
            saved_count += 1
            saved_pdf += 1
            if save_json:
                json_path = target_dir / artifact["output_pdf_name"].replace(".pdf", ".results.json")
                json_path.write_bytes(artifact["results_json_bytes"])
                saved_count += 1
                saved_json += 1
        if save_zip and batch_zip_bytes and batch_zip_name:
            (target_dir / batch_zip_name).write_bytes(batch_zip_bytes)
            saved_count += 1
            saved_zip += 1
        detail = f"PDF {saved_pdf}"
        if save_json:
            detail += f", JSON {saved_json}"
        if save_zip:
            detail += f", ZIP {saved_zip}"
        return True, f"已保存到: {target_dir}（共 {saved_count} 个文件；{detail}）"
    except Exception as exc:
        return False, f"保存失败: {exc}"


def _browser_download_payload() -> tuple[bytes, str, str]:
    artifacts = st.session_state.get("artifacts", [])
    if not artifacts:
        return b"", "", ""
    zip_bytes = st.session_state.get("batch_zip_bytes", b"")
    zip_name = st.session_state.get("batch_zip_name", "")
    if len(artifacts) > 1 and zip_bytes and zip_name:
        return zip_bytes, str(zip_name), "application/zip"
    first = artifacts[0]
    return (
        bytes(first.get("output_pdf_bytes", b"")),
        str(first.get("output_pdf_name", "translated.pdf")),
        "application/pdf",
    )


def _render_browser_auto_download() -> None:
    run_id = str(st.session_state.get("browser_download_run_id", ""))
    if not run_id or st.session_state.get("browser_download_rendered_run_id") == run_id:
        return
    data, filename, mime = _browser_download_payload()
    if not data or not filename:
        return
    # Browser auto-downloads can be blocked by site settings. The normal Streamlit
    # download buttons remain the reliable fallback immediately below.
    b64 = base64.b64encode(data).decode("ascii")
    safe_filename = html.escape(filename, quote=True)
    safe_mime = html.escape(mime, quote=True)
    components.html(
        f"""
        <html>
          <body>
            <a id="autopdf-download" download="{safe_filename}" href="data:{safe_mime};base64,{b64}"></a>
            <script>
              const link = document.getElementById("autopdf-download");
              if (link) {{
                setTimeout(() => link.click(), 250);
              }}
            </script>
          </body>
        </html>
        """,
        height=0,
    )
    st.session_state["browser_download_rendered_run_id"] = run_id


def _append_translation_history(entries: list[dict[str, str]]) -> None:
    if not entries:
        return
    for item in entries:
        if not str(item.get("id", "")).strip():
            item["id"] = str(uuid.uuid4())
    history = list(st.session_state.get("translation_history", []))
    history.extend(entries)
    history = history[-300:]
    st.session_state["translation_history"] = history
    try:
        _save_translation_history(history)
    except Exception as exc:
        st.warning(f"Failed to persist translation history: {exc}")


def _format_task_duration(seconds: float) -> str:
    total = max(0.0, float(seconds))
    if total < 60:
        return f"{total:.1f} 秒"
    minutes = int(total // 60)
    remain = total - minutes * 60
    return f"{minutes} 分 {remain:.1f} 秒"


def _build_history_entries(
    artifacts: list[dict[str, Any]],
    *,
    export_output_dir: str,
    mode_label: str,
    provider_group: str,
    provider_config: dict[str, str],
    auto_saved: bool,
    task_duration_seconds: float,
) -> list[dict[str, str]]:
    now_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mode_text = MODE_HISTORY_LABEL.get(mode_label, mode_label)
    duration_text = _format_task_duration(task_duration_seconds)
    output_dir = Path(export_output_dir).expanduser() if export_output_dir.strip() else None
    entries: list[dict[str, str]] = []
    for artifact in artifacts:
        source_file = str(artifact.get("filename", ""))
        translated_file = str(artifact.get("output_pdf_name", ""))
        page_stats = artifact.get("page_stats", [])
        total_comments = 0
        translation_errors = 0
        if isinstance(page_stats, list):
            for row in page_stats:
                if isinstance(row, dict):
                    try:
                        total_comments += int(row.get("Comments", 0))
                    except Exception:
                        pass
                    try:
                        translation_errors += int(row.get("Translation errors", 0))
                    except Exception:
                        pass
        if output_dir is not None and auto_saved:
            translated_path = str((output_dir / translated_file).resolve())
            result_text = "成功"
        else:
            translated_path = "(未自动保存)"
            result_text = "成功(未自动保存)"
        if mode_text == "Vision" and total_comments == 0:
            result_text = "失败（未识别到意见）"
        elif translation_errors:
            result_text = "失败（部分未翻译）"
        entries.append(
            {
                "id": str(uuid.uuid4()),
                "translated_at": now_text,
                "source_file": source_file,
                "source_path": f"上传文件/{source_file}",
                "translated_file": translated_file,
                "translated_path": translated_path,
                "mode": mode_text,
                "provider_group": provider_group,
                "provider_name": str(provider_config.get("provider_name", "")),
                "provider_model": str(provider_config.get("model", "")),
                "result": result_text,
                "task_duration": duration_text,
            }
        )
    return entries


def _artifact_failure_reason(artifacts: list[dict[str, Any]], mode_label: str) -> str:
    total_comments = 0
    translation_errors = 0
    for artifact in artifacts:
        page_stats = artifact.get("page_stats", [])
        if not isinstance(page_stats, list):
            continue
        for row in page_stats:
            if not isinstance(row, dict):
                continue
            try:
                total_comments += int(row.get("Comments", 0))
            except Exception:
                pass
            try:
                translation_errors += int(row.get("Translation errors", 0))
            except Exception:
                pass
    if MODE_HISTORY_LABEL.get(mode_label, mode_label) == "Vision" and total_comments == 0:
        return "未识别到意见"
    if translation_errors:
        return "部分未翻译"
    return ""


def _pick_index(current: str, options: list[str]) -> int:
    if current in options:
        return options.index(current)
    return 0


def _normalize_bearer_api_key(value: str) -> str:
    key = str(value or "").strip()
    if key.lower().startswith("bearer "):
        key = key[7:].strip()
    return key


def _is_minimax_config(config: dict[str, str]) -> bool:
    provider = config.get("provider_name", "").lower()
    base = config.get("base_url", "").lower()
    return "minimax" in provider or "minimax.io" in base or "minimaxi.com" in base


def _models_urls(base_url: str) -> list[str]:
    base = base_url.strip().rstrip("/")
    if not base:
        return []
    candidates = [f"{base}/models"]
    if "/v1" not in base:
        candidates.append(f"{base}/v1/models")
    dedup: list[str] = []
    for item in candidates:
        if item not in dedup:
            dedup.append(item)
    return dedup


def _check_connection(config: dict[str, str]) -> tuple[bool, str]:
    provider_impl = config.get("provider_impl", "openai-compatible")
    if provider_impl == "mock":
        return True, "Mock provider is ready (offline test mode)."
    if requests is None:
        return False, "requests is not installed, cannot run connection check."

    base_url = config.get("base_url", "").strip()
    if not base_url:
        return False, "Base URL is required."

    headers = {"Content-Type": "application/json"}
    api_key = _normalize_bearer_api_key(config.get("api_key", ""))
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"

    if _is_minimax_config(config):
        try:
            base = base_url.rstrip("/")
            response = requests.post(
                f"{base}/chat/completions",
                headers=headers,
                json={
                    "model": config.get("model", "").strip() or "MiniMax-M2.7",
                    "messages": [{"role": "user", "content": "Return only: ok"}],
                    "temperature": 0.0,
                    "max_tokens": 8,
                    "reasoning_split": True,
                },
                timeout=20,
            )
            if response.status_code == 200:
                return True, "Connected successfully via MiniMax chat completions."
            return False, f"MiniMax chat check failed: HTTP {response.status_code}. {_extract_http_body(response)}"
        except Exception as exc:
            return False, f"MiniMax chat check failed: {exc}"

    errors: list[str] = []
    for url in _models_urls(base_url):
        try:
            response = requests.get(url, headers=headers, timeout=12)
            if response.status_code != 200:
                errors.append(f"{url} -> HTTP {response.status_code}")
                continue
            payload = response.json()
            models = [str(item.get("id", "")) for item in payload.get("data", []) if isinstance(item, dict)]
            model = config.get("model", "").strip()
            if model and models and model not in models:
                return True, f"Connected, but model '{model}' was not listed by server."
            return True, f"Connected successfully via {url}."
        except Exception as exc:
            errors.append(f"{url} -> {exc}")
    return False, "Connection failed. " + " | ".join(errors[:2])


def _extract_http_body(response: Any) -> str:
    try:
        payload = response.json()
        return json.dumps(payload, ensure_ascii=False)[:800]
    except Exception:
        return str(getattr(response, "text", "") or "")[:800]


def _build_config_from_inputs(
    preset_map: dict[str, dict[str, str]],
    provider_name: str,
    base_url: str,
    api_key: str,
    model: str,
    prompt_hint: str,
) -> dict[str, str]:
    preset = preset_map[provider_name]
    return {
        "provider_name": provider_name,
        "base_url": base_url.strip(),
        "api_key": _normalize_bearer_api_key(api_key),
        "model": model.strip(),
        "prompt_hint": prompt_hint.strip(),
        "provider_impl": preset["provider_impl"],
    }


def _apply_runtime_provider(config: dict[str, str]) -> str:
    provider_impl = config.get("provider_impl", "openai-compatible")
    if provider_impl == "mock":
        os.environ.pop("OPENAI_API_KEY", None)
        os.environ.pop("OPENAI_BASE_URL", None)
        os.environ.pop("AUTOPDFTRANSLATOR_MODEL", None)
        os.environ.pop("AUTOPDFTRANSLATOR_PROMPT_HINT", None)
        _clear_runtime_memory_env()
        os.environ.pop("AUTOPDFTRANSLATOR_ALLOW_EMPTY_KEY", None)
        return "mock"

    os.environ["OPENAI_BASE_URL"] = config.get("base_url", "").strip()
    os.environ["AUTOPDFTRANSLATOR_MODEL"] = config.get("model", "").strip() or "gpt-4.1-mini"
    prompt_hint = _build_effective_prompt_hint(config.get("prompt_hint", ""))
    if prompt_hint:
        os.environ["AUTOPDFTRANSLATOR_PROMPT_HINT"] = prompt_hint
    else:
        os.environ.pop("AUTOPDFTRANSLATOR_PROMPT_HINT", None)
    if bool(st.session_state.get("corpus_enabled", True)):
        memory = st.session_state.get("translation_memory", {})
        if isinstance(memory, dict) and (memory.get("terms") or memory.get("style_rules") or memory.get("examples")):
            _apply_runtime_memory(memory)
        else:
            _clear_runtime_memory_env()
    else:
        _clear_runtime_memory_env()

    api_key = config.get("api_key", "").strip()
    api_key = _normalize_bearer_api_key(api_key)
    if api_key:
        os.environ["OPENAI_API_KEY"] = api_key
        os.environ.pop("AUTOPDFTRANSLATOR_ALLOW_EMPTY_KEY", None)
    else:
        os.environ.pop("OPENAI_API_KEY", None)
        os.environ["AUTOPDFTRANSLATOR_ALLOW_EMPTY_KEY"] = "1"
    return "openai-compatible"


PROGRESS_STAGE_LABELS = [
    ("extracting", "提取意见"),
    ("translating", "AI翻译"),
    ("writing", "写入PDF"),
    ("done", "完成"),
]


def _stage_index(stage: str, percent: int) -> int:
    if percent >= 100 or stage == "done":
        return 3
    stage_order = {name: idx for idx, (name, _) in enumerate(PROGRESS_STAGE_LABELS)}
    return stage_order.get(stage, 0)


def _render_stage_progress(progress_slot: Any) -> None:
    percent = max(0, min(100, int(st.session_state.get("task_progress_percent", 0) or 0)))
    text = html.escape(str(st.session_state.get("task_progress_text", "任务未开始")))
    stage = str(st.session_state.get("task_progress_stage", "idle"))
    state_class = "is-static" if stage in {"idle", "done"} or percent <= 0 or percent >= 100 else "is-running"
    current_idx = _stage_index(stage, percent)
    checks = []
    labels = []
    for idx, (_, label) in enumerate(PROGRESS_STAGE_LABELS):
        check_class = "check"
        label_class = ""
        if idx < current_idx or percent >= 100:
            check_class += " is-complete"
        elif idx == current_idx and percent > 0:
            check_class += " is-current"
            label_class = ' class="is-active"'
        checks.append(f'<div class="{check_class}"></div>')
        labels.append(f"<span{label_class}>{html.escape(label)}</span>")

    confetti_html = ""
    confetti_run_id = str(st.session_state.get("confetti_run_id", ""))
    confetti_rendered_run_id = str(st.session_state.get("confetti_rendered_run_id", ""))
    if stage == "done" and percent >= 100 and confetti_run_id and confetti_run_id != confetti_rendered_run_id:
        confetti_html = _build_confetti_html()
        st.session_state["confetti_rendered_run_id"] = confetti_run_id
    failure_html = ""
    failure_run_id = str(st.session_state.get("failure_animation_run_id", ""))
    failure_rendered_run_id = str(st.session_state.get("failure_animation_rendered_run_id", ""))
    if stage == "failed" and failure_run_id and failure_run_id != failure_rendered_run_id:
        failure_html = _build_failure_animation_html()
        st.session_state["failure_animation_rendered_run_id"] = failure_run_id
    extra_state_class = "is-failed" if stage == "failed" else ("is-idle" if stage == "idle" or percent <= 0 else "")
    progress_slot.markdown(
        f"""
<div class="stage-progress {state_class} {extra_state_class}">
    <div class="stage-progress__meta">
        <span>{text}</span>
    </div>
    <div class="stage-progress__row">
        <div class="stage-progress__main">
            <div class="loader" style="--progress-width: {percent}%;">
                <div class="bar"></div>
                <div class="check-bar-container">
                    {''.join(checks)}
                </div>
            </div>
            <div class="stage-progress__labels">
                {''.join(labels)}
            </div>
        </div>
        <div class="typewriter" aria-hidden="true">
            <div class="slide"><i></i></div>
            <div class="paper"></div>
            <div class="keyboard"></div>
        </div>
        <span class="stage-progress__percent">{percent}%</span>
    </div>
</div>
{confetti_html}
{failure_html}
        """,
        unsafe_allow_html=True,
    )


def _uploaded_files_signature(uploaded_pdfs: list[Any] | None) -> str:
    if not uploaded_pdfs:
        return ""
    parts: list[str] = []
    for uploaded in uploaded_pdfs:
        name = str(getattr(uploaded, "name", ""))
        size = int(getattr(uploaded, "size", 0) or 0)
        parts.append(f"{name}:{size}")
    return "|".join(parts)


def _reset_run_state_for_upload_change() -> None:
    st.session_state["task_progress_percent"] = 0
    st.session_state["task_progress_text"] = "任务未开始"
    st.session_state["task_progress_stage"] = "idle"
    st.session_state["confetti_run_id"] = ""
    st.session_state["confetti_rendered_run_id"] = ""
    st.session_state["failure_animation_run_id"] = ""
    st.session_state["failure_animation_rendered_run_id"] = ""
    st.session_state["translation_running"] = False
    st.session_state["translation_cancel_requested"] = False
    st.session_state["run_message"] = ""
    st.session_state["run_error"] = ""
    st.session_state["artifacts"] = []
    st.session_state["batch_zip_bytes"] = b""
    st.session_state["batch_zip_name"] = ""
    st.session_state["save_path_message"] = ""
    st.session_state["preflight_reports"] = []
    st.session_state["preflight_key"] = ""


def _request_translation_cancel() -> None:
    st.session_state["translation_cancel_requested"] = True
    st.session_state["task_progress_text"] = "正在终止：等待当前步骤结束"


def _raise_if_translation_cancelled() -> None:
    if bool(st.session_state.get("translation_cancel_requested", False)):
        raise TranslationCancelled("用户终止了当前翻译任务。")


def _build_confetti_html() -> str:
    colors = ["#22c55e", "#2563eb", "#f59e0b", "#ef4444", "#a855f7", "#06b6d4", "#f97316"]
    pieces: list[str] = []
    for idx in range(96):
        angle = (idx * 137) % 360
        spread = 170 + (idx % 12) * 34
        x = round(math.cos(math.radians(angle)) * spread, 1)
        y = round(math.sin(math.radians(angle)) * spread * 0.68 - 90, 1)
        drift = x + ((idx % 9) - 4) * 48
        fall = 620 + (idx % 13) * 34
        color = colors[idx % len(colors)]
        delay = round((idx % 12) * 0.026, 3)
        r1 = 160 + idx * 23
        r2 = 540 + idx * 31
        pieces.append(
            (
                '<span class="confetti-piece" '
                f'style="--confetti-color:{color};--confetti-x:{x}px;--confetti-y:{y}px;'
                f'--confetti-x2:{drift}px;--confetti-fall:{fall}px;'
                f'--confetti-r1:{r1}deg;--confetti-r2:{r2}deg;'
                f'--confetti-delay:{delay}s;"></span>'
            )
        )
    return f'<div class="confetti-burst" aria-hidden="true">{"".join(pieces)}</div>'


def _build_failure_animation_html() -> str:
    sparks = "".join('<span class="fault-spark"></span>' for _ in range(4))
    return f'<div class="fault-burst" aria-hidden="true"><div class="fault-machine">{sparks}</div></div>'


def _build_task_progress_callback(progress_slot: Any, file_index: int, file_total: int) -> Any:
    file_start = int((file_index - 1) / max(file_total, 1) * 100)
    file_span = max(1, int(100 / max(file_total, 1)))
    stage_ranges = {
        "extracting": (0, 45, "提取英文意见"),
        "translating": (45, 82, "AI批量翻译"),
        "writing": (82, 100, "写入PDF文本框"),
    }

    def callback(event: dict[str, Any]) -> None:
        _raise_if_translation_cancelled()
        stage = str(event.get("stage", ""))
        start, end, label = stage_ranges.get(stage, (0, 100, "处理中"))
        total = max(0, int(event.get("pages_total", 0) or 0))
        done = max(0, int(event.get("pages_done", 0) or 0))
        remaining = max(0, total - done)
        ratio = (done / total) if total else 0.0
        within_file = start + int((end - start) * ratio)
        pct = min(99, file_start + int(file_span * within_file / 100))
        if stage == "translating" and str(event.get("phase", "")) == "batch_start":
            comments_total = int(event.get("comments_total", 0) or 0)
            text = f"{label}中：已提交一次性请求，覆盖 {total} 页 / {comments_total} 条意见；等待AI返回"
        elif stage == "translating" and str(event.get("phase", "")) == "batch_error":
            text = f"{label}失败：AI文本翻译未返回有效结果，已保留原文。请查看日志。"
        else:
            text = f"{label}中：已完成 {done}/{total} 页，剩余 {remaining} 页"
        if file_total > 1:
            text = f"文件 {file_index}/{file_total} - {text}"
        st.session_state["task_progress_percent"] = pct
        st.session_state["task_progress_text"] = text
        st.session_state["task_progress_stage"] = stage
        _render_stage_progress(progress_slot)

    return callback


@st.cache_data(show_spinner=False)
def _classify_pdf_bytes(filename: str, pdf_bytes: bytes) -> dict[str, Any]:
    import fitz

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    try:
        page_count = doc.page_count
        annotation_pages = 0
        annotation_items = 0
        text_candidate_items = 0
        text_pages = 0
        total_text_chars = 0
        image_pages = 0

        for page in doc:
            page_text = page.get_text("text") or ""
            text_chars = len(page_text.strip())
            total_text_chars += text_chars
            if text_chars >= 30:
                text_pages += 1

            has_image = False
            try:
                has_image = bool(page.get_images(full=True))
            except Exception:
                has_image = False
            if has_image:
                image_pages += 1

            try:
                from autopdftranslator.extractors import PDFTextExtractor

                text_candidate_items += len(PDFTextExtractor().extract_page_comments(page, page.number).comments)
            except Exception:
                pass

            page_annotation_count = 0
            try:
                annotations = list(page.annots() or [])
            except Exception:
                annotations = []
            for annot in annotations:
                try:
                    type_name = str(annot.type[1])
                except Exception:
                    type_name = ""
                if type_name != "FreeText":
                    continue
                info = getattr(annot, "info", {}) or {}
                subject = str(info.get("subject", "") or "")
                title = str(info.get("title", "") or "")
                if APP_NAME.lower() in subject.lower() or APP_NAME.lower() in title.lower():
                    continue
                content = str(info.get("content", "") or "").strip()
                if re.search(r"[A-Za-z]", content):
                    page_annotation_count += 1
            if page_annotation_count:
                annotation_pages += 1
                annotation_items += page_annotation_count

        avg_text_chars = int(total_text_chars / max(page_count, 1))
        image_ratio = image_pages / max(page_count, 1)
        annotation_ratio = annotation_pages / max(page_count, 1)
        text_ratio = text_pages / max(page_count, 1)

        if annotation_items:
            pdf_type = "原生批注型 PDF"
            recommended_mode = "text extraction only"
            reason = f"检测到 {annotation_items} 条英文 FreeText 批注，直接提取最快最稳。"
        elif image_ratio >= 0.6 and avg_text_chars < 120:
            pdf_type = "图片/手写混合型 PDF"
            recommended_mode = "hybrid"
            reason = "页面以图片为主且文本层较少，可能包含手写批注；建议用 hybrid 触发视觉识别兜底，手写较多时可手动切到 vision only。"
        elif text_ratio >= 0.6 and avg_text_chars >= 80:
            pdf_type = "可提取文本型 PDF"
            recommended_mode = "text extraction only"
            reason = "多数页面有可提取文本层，优先使用文本模式以降低成本并提升速度。"
        elif image_ratio >= 0.6 and avg_text_chars < 40:
            pdf_type = "扫描/图片型 PDF"
            recommended_mode = "vision only"
            reason = "多数页面像图片或扫描件，文本层很少，需要视觉识别。"
        else:
            pdf_type = "混合型 PDF"
            recommended_mode = "hybrid"
            reason = "部分页面有文本层，部分页面可能需要视觉兜底。"

        return {
            "filename": filename,
            "pages": page_count,
            "pdf_type": pdf_type,
            "recommended_mode": recommended_mode,
            "reason": reason,
            "annotation_items": annotation_items,
            "annotation_pages": annotation_pages,
            "text_candidate_items": text_candidate_items,
            "text_pages": text_pages,
            "image_pages": image_pages,
            "avg_text_chars": avg_text_chars,
            "annotation_ratio": annotation_ratio,
            "text_ratio": text_ratio,
            "image_ratio": image_ratio,
        }
    finally:
        doc.close()


def _classify_uploaded_pdfs(uploaded_pdfs: list[Any]) -> list[dict[str, Any]]:
    reports: list[dict[str, Any]] = []
    for uploaded in uploaded_pdfs:
        try:
            reports.append(_classify_pdf_bytes(uploaded.name, uploaded.getvalue()))
        except Exception as exc:
            reports.append(
                {
                    "filename": getattr(uploaded, "name", "unknown.pdf"),
                    "pages": 0,
                    "pdf_type": "无法判定",
                    "recommended_mode": "hybrid",
                    "reason": f"读取 PDF 结构失败：{exc}",
                    "annotation_items": 0,
                    "annotation_pages": 0,
                    "text_candidate_items": 0,
                    "text_pages": 0,
                    "image_pages": 0,
                    "avg_text_chars": 0,
                }
            )
    return reports


def _choose_overall_recommended_mode(reports: list[dict[str, Any]]) -> str:
    modes = [str(report.get("recommended_mode", "")) for report in reports]
    if not modes:
        return "hybrid"
    if "vision only" in modes and any(mode != "vision only" for mode in modes):
        return "hybrid"
    if "hybrid" in modes:
        return "hybrid"
    if "vision only" in modes:
        return "vision only"
    return "text extraction only"


def _render_pdf_type_reports(reports: list[dict[str, Any]], *, vision_supported: bool) -> str:
    if not reports:
        return "hybrid"
    recommended_mode = _choose_overall_recommended_mode(reports)
    effective_recommendation = recommended_mode
    if not vision_supported and recommended_mode in {"vision only", "hybrid"}:
        effective_recommendation = "text extraction only"

    with st.expander("PDF类型判定 / Mode Recommendation", expanded=True):
        rows = [
            {
                "文件": report.get("filename", ""),
                "类型": report.get("pdf_type", ""),
                "推荐模式": report.get("recommended_mode", ""),
                "页数": report.get("pages", 0),
                "原生批注": report.get("annotation_items", 0),
                "文本候选意见": report.get("text_candidate_items", 0),
                "文本页": report.get("text_pages", 0),
                "图片页": report.get("image_pages", 0),
                "依据": report.get("reason", ""),
            }
            for report in reports
        ]
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
        if effective_recommendation != recommended_mode:
            st.warning(
                f"检测建议为 {recommended_mode}，但当前模型配置不支持视觉，因此界面将限制为 text extraction only。"
            )
        else:
            st.info(f"建议翻译模式：{effective_recommendation}。可手动切换。")
        if vision_supported and any(
            str(report.get("pdf_type", "")) in {"图片/手写混合型 PDF", "扫描/图片型 PDF"}
            for report in reports
        ):
            st.caption(
                "手写批注较多时，建议使用 Gemini 或 OpenAI 视觉模型；Kimi 等推理模型可能会把输出预算消耗在 reasoning，导致识别为空或耗时偏长。"
            )
    return effective_recommendation


def _validate_runtime_config(config: dict[str, str], group: str) -> str | None:
    provider_impl = config.get("provider_impl", "openai-compatible")
    if provider_impl == "mock":
        return None
    if not config.get("base_url", "").strip():
        return f"{group}: Base URL is required."
    if group == "API Providers" and not config.get("api_key", "").strip():
        return "API Providers: API key is required."
    if not config.get("model", "").strip():
        return f"{group}: Model is required."
    return None


def _is_mode_requires_vision(mode_label: str) -> bool:
    return mode_label.strip().lower() in {"vision only", "hybrid"}


def _model_looks_vision_capable(model_name: str) -> bool:
    model = model_name.strip().lower()
    if not model:
        return False
    keywords = [
        "vision",
        "vl",
        "4v",
        "gpt-4o",
        "gpt-4.1",
        "omni",
        "gemini",
        "kimi",
        "k2.5",
        "llava",
        "qwen-vl",
        "glm-4v",
        "internvl",
        "minicpm-v",
    ]
    return any(key in model for key in keywords)


def _is_vision_supported(config: dict[str, str], group: str) -> tuple[bool, str]:
    provider_impl = config.get("provider_impl", "openai-compatible")
    provider_name = config.get("provider_name", "").strip()
    model_name = config.get("model", "").strip()
    base_url = config.get("base_url", "").strip().lower()

    if provider_impl == "mock":
        return False, "Mock provider 不执行真实视觉提取。"

    if group == "Local Models" or "localhost" in base_url or "127.0.0.1" in base_url:
        return True, "本地模型无法仅凭模型名可靠判断视觉能力；程序允许视觉模式，实际是否可用由本地模型/服务端决定。"

    if group == "API Providers":
        if provider_name == "Moonshot":
            ok = _model_looks_vision_capable(model_name)
            return (ok, "" if ok else "Moonshot 需要选择支持视觉的模型（建议包含 kimi/k2.5/vl/vision 等标识）。")
        if provider_name == "DeepSeek":
            ok = _model_looks_vision_capable(model_name)
            return (ok, "" if ok else "DeepSeek 需要选择支持视觉的模型（如包含 vl/vision 的模型）。")
        if provider_name == "OpenAI":
            ok = _model_looks_vision_capable(model_name)
            return (ok, "" if ok else "当前 OpenAI 模型看起来不支持视觉输入。")
        if provider_name == "Gemini (OpenAI-Compatible)":
            return True, ""
        if provider_name == "Qwen Compatible":
            ok = _model_looks_vision_capable(model_name)
            return (ok, "" if ok else "Qwen 需要选择视觉模型（如包含 vl/vision 的模型）。")

    ok = _model_looks_vision_capable(model_name)
    return (ok, "" if ok else "当前模型看起来不支持视觉输入，请切换到视觉模型。")


def _refresh_detection_cache() -> None:
    _analyze_uploaded_pdf.clear()


@st.cache_data(show_spinner=False)
def _analyze_uploaded_pdf(
    pdf_bytes: bytes,
    filename: str,
    mode: str,
    typed_text_threshold: int,
    prompt_overhead_tokens: int,
    vision_input_tokens_per_page: int,
    vision_output_tokens_per_page: int,
    input_cost_per_1m: float,
    output_cost_per_1m: float,
) -> dict[str, Any]:
    with tempfile.TemporaryDirectory() as temp_dir:
        input_path = Path(temp_dir) / filename
        input_path.write_bytes(pdf_bytes)
        return analyze_pdf(
            input_path,
            mode=mode,
            typed_text_threshold=typed_text_threshold,
            prompt_overhead_tokens=prompt_overhead_tokens,
            vision_input_tokens_per_page=vision_input_tokens_per_page,
            vision_output_tokens_per_page=vision_output_tokens_per_page,
            input_cost_per_1m=input_cost_per_1m,
            output_cost_per_1m=output_cost_per_1m,
        )


def _build_preflight_key(
    uploaded_pdfs: list[Any],
    *,
    mode: str,
    typed_text_threshold: int,
    prompt_overhead_tokens: int,
    vision_input_tokens_per_page: int,
    vision_output_tokens_per_page: int,
    input_cost_per_1m: float,
    output_cost_per_1m: float,
) -> str:
    file_signatures: list[str] = []
    for uploaded in uploaded_pdfs:
        size_val = getattr(uploaded, "size", None)
        if size_val is None:
            size_val = len(uploaded.getvalue())
        file_signatures.append(f"{uploaded.name}:{int(size_val)}")

    payload = {
        "files": file_signatures,
        "mode": mode,
        "typed_text_threshold": int(typed_text_threshold),
        "prompt_overhead_tokens": int(prompt_overhead_tokens),
        "vision_input_tokens_per_page": int(vision_input_tokens_per_page),
        "vision_output_tokens_per_page": int(vision_output_tokens_per_page),
        "input_cost_per_1m": float(input_cost_per_1m),
        "output_cost_per_1m": float(output_cost_per_1m),
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)


def _render_preflight_reports(reports: list[dict[str, Any]], usd_to_cny_rate: float) -> None:
    if not reports:
        return

    total_pages = sum(int(report["totals"]["pages"]) for report in reports)
    total_comments = sum(int(report["totals"]["text_comments"]) for report in reports)
    total_tokens = sum(int(report["totals"]["total_tokens"]) for report in reports)
    total_cost_usd = sum(float(report["totals"]["estimated_cost_usd"]) for report in reports)
    total_cost_cny = total_cost_usd * usd_to_cny_rate

    with st.expander("消耗预测 / Pre-run Estimate", expanded=False):
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("PDFs", len(reports))
        c2.metric("Pages", total_pages)
        c3.metric("Detected Comments", total_comments)
        c4.metric("Est. Cost (RMB)", f"¥{total_cost_cny:.2f}")
        st.caption(
            f"Estimated token usage: {total_tokens:,} tokens. About ${total_cost_usd:.4f} USD. "
            "This is a rough estimate and can differ from actual billing."
        )

        for report in reports:
            totals = report["totals"]
            file_cost_usd = float(totals["estimated_cost_usd"])
            file_cost_cny = file_cost_usd * usd_to_cny_rate
            with st.expander(
                f"{report['filename']} · {totals['pages']} pages · ¥{file_cost_cny:.2f} (${file_cost_usd:.4f})"
            ):
                stat_a, stat_b, stat_c = st.columns(3)
                stat_a.metric("Input tokens", f"{int(totals['input_tokens']):,}")
                stat_b.metric("Output tokens", f"{int(totals['output_tokens']):,}")
                stat_c.metric("Predicted vision pages", int(totals["predicted_vision_pages"]))

                table_rows: list[dict[str, object]] = []
                for page in report["pages"]:
                    table_rows.append(
                        {
                            "Page": page["page_index"],
                            "Text chars": page["text_chars"],
                            "Detected comments": len(page["text_comments"]),
                            "Predicted vision": page["will_use_vision"],
                        }
                    )
                st.dataframe(table_rows, width="stretch", hide_index=True)

                st.markdown("**Detected comments by page**")
                for page in report["pages"]:
                    label = f"Page {page['page_index']} ({len(page['text_comments'])} comments)"
                    with st.expander(label):
                        comments = page["text_comments"]
                        if comments:
                            preview = comments[:20]
                            st.code("\n".join(f"- {comment}" for comment in preview), language="text")
                            if len(comments) > len(preview):
                                st.caption(f"Showing first {len(preview)} of {len(comments)} comments.")
                        else:
                            st.caption("No text-layer comments detected on this page.")
                        if page["will_use_vision"]:
                            st.caption("Current mode predicts vision will be used on this page.")


def _render_outputs() -> None:
    if st.session_state["run_error"]:
        st.error(st.session_state["run_error"])
    elif st.session_state["run_message"]:
        st.success(st.session_state["run_message"])

    artifacts = st.session_state["artifacts"]
    if not artifacts:
        return

    total_pages = sum(len(item["page_stats"]) for item in artifacts)
    total_comments = sum(sum(int(row["Comments"]) for row in item["page_stats"]) for item in artifacts)
    total_vision_pages = sum(
        sum(1 for row in item["page_stats"] if row["Vision used"])
        for item in artifacts
    )
    total_text_pages = max(0, total_pages - total_vision_pages)
    vision_ratio = (total_vision_pages / total_pages * 100.0) if total_pages else 0.0
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Translated PDFs", len(artifacts))
    m2.metric("Pages", total_pages)
    m3.metric("Comments", total_comments)
    m4.metric("Vision Pages", total_vision_pages)
    m5.metric("Text Pages", total_text_pages)
    st.caption(
        f"Execution ratio in this run: Vision {total_vision_pages}/{total_pages} ({vision_ratio:.1f}%), "
        f"Text {total_text_pages}/{total_pages} ({100.0 - vision_ratio:.1f}%)."
    )
    _render_browser_auto_download()

    if _server_path_saving_enabled():
        save_col1, save_col2 = st.columns([3, 1], gap="small")
        with save_col1:
            st.caption(f"指定保存路径: {st.session_state.get('export_output_dir', '')}")
        with save_col2:
            manual_save_clicked = st.button("保存到指定路径", width="stretch", key="manual_save_to_path")

        if manual_save_clicked:
            ok, message = _save_artifacts_to_path(
                artifacts,
                str(st.session_state.get("export_output_dir", "")),
                st.session_state.get("batch_zip_bytes", b""),
                st.session_state.get("batch_zip_name", ""),
                save_json=True,
                save_zip=True,
            )
            st.session_state["save_path_message"] = message
            if ok:
                _persist_ui_settings()
                st.success(message)
            else:
                st.error(message)
        elif st.session_state.get("save_path_message"):
            st.caption(str(st.session_state["save_path_message"]))
    else:
        st.info("当前运行在 Streamlit Cloud，服务器保存路径不可直接访问；请使用下方浏览器下载按钮保存文件。")

    if st.session_state["batch_zip_bytes"]:
        st.download_button(
            label="Download All (ZIP)",
            data=st.session_state["batch_zip_bytes"],
            file_name=st.session_state["batch_zip_name"],
            mime="application/zip",
            type="primary",
            width="stretch",
        )

    for artifact in artifacts:
        with st.expander(f"Result · {artifact['filename']}"):
            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    label="Download translated PDF",
                    data=artifact["output_pdf_bytes"],
                    file_name=artifact["output_pdf_name"],
                    mime="application/pdf",
                    width="stretch",
                )
            with dl2:
                st.download_button(
                    label="Download results JSON",
                    data=artifact["results_json_bytes"],
                    file_name=artifact["output_pdf_name"].replace(".pdf", ".results.json"),
                    mime="application/json",
                    width="stretch",
                )

            t_stats, t_logs = st.tabs(["Page Stats", "Logs"])
            with t_stats:
                st.dataframe(artifact["page_stats"], width="stretch", hide_index=True)
            with t_logs:
                if artifact["log_lines"]:
                    st.code("\n".join(artifact["log_lines"]), language="text")
                else:
                    st.caption("No logs were emitted.")


def _render_translation_history() -> None:
    history = st.session_state.get("translation_history", [])
    with st.expander(f"翻译历史 ({len(history)})", expanded=False):
        _render_translation_history_body(history)


def _render_translation_history_body(history: list[dict[str, str]]) -> None:
    if not history:
        st.caption("暂无历史记录。")
        return

    for item in history:
        if not str(item.get("id", "")).strip():
            item["id"] = str(uuid.uuid4())

    if bool(st.session_state.get("history_clear_search_requested", False)):
        st.session_state["history_search_query"] = ""
        st.session_state["history_clear_search_requested"] = False

    search_col, clear_col = st.columns([6, 1], gap="small")
    with search_col:
        search_query = str(
            st.text_input(
                "搜索翻译历史",
                key="history_search_query",
                placeholder="输入文件名、翻译模式、结果、日期等关键字",
            )
            or ""
        ).strip()
    with clear_col:
        st.write("")
        clear_search_clicked = st.button(
            "清空搜索",
            key="history_clear_search_btn",
            width="stretch",
            disabled=not search_query,
        )
    search_lower = search_query.lower()

    if search_lower:
        filtered_history = [
            item
            for item in history
            if search_lower
            in " ".join(
                [
                    str(item.get("translated_at", "")),
                    str(item.get("source_file", "")),
                    str(item.get("translated_file", "")),
                    str(item.get("mode", "")),
                    str(item.get("provider_group", "")),
                    str(item.get("provider_name", "")),
                    str(item.get("provider_model", "")),
                    str(item.get("result", "")),
                    str(item.get("task_duration", "")),
                ]
            ).lower()
        ]
    else:
        filtered_history = list(history)

    if search_lower:
        if not filtered_history:
            st.caption("未找到匹配记录。")
        else:
            st.caption(f"检索到 {len(filtered_history)} 条记录。")

    rows: list[dict[str, Any]] = []
    for item in reversed(filtered_history):
        item_id = str(item.get("id", ""))
        rows.append(
            {
                "_id": item_id,
                "翻译日期": str(item.get("translated_at", "")),
                "文件名": str(item.get("source_file", "")),
                "翻译后文件名": str(item.get("translated_file", "")),
                "翻译模式": str(item.get("mode", "")),
                "运行方式": str(item.get("provider_group", "")),
                "Provider": str(item.get("provider_name", "")),
                "Model": str(item.get("provider_model", "")),
                "任务耗时": str(item.get("task_duration", "")),
                "翻译结果": str(item.get("result", "")),
            }
        )

    delete_mode = bool(st.session_state.get("history_delete_mode", False))
    selected_ids: list[str] = []

    if rows:
        with st.container(height=360, border=True):
            if delete_mode:
                edit_rows = []
                for row in rows:
                    row_copy = dict(row)
                    row_copy["删除"] = False
                    edit_rows.append(row_copy)
                df_edit = pd.DataFrame(edit_rows)
                edited = st.data_editor(
                    df_edit,
                    width="stretch",
                    hide_index=True,
                    column_config={
                        "_id": None,
                        "删除": st.column_config.CheckboxColumn("删除", help="勾选要删除的记录"),
                    },
                    disabled=[c for c in df_edit.columns if c != "删除"],
                )
                selected_ids = [
                    str(row["_id"])
                    for _, row in edited.iterrows()
                    if bool(row.get("删除", False))
                ]
            else:
                st.dataframe(
                    [{k: v for k, v in row.items() if k != "_id"} for row in rows],
                    width="stretch",
                    hide_index=True,
                )
    else:
        st.caption("当前无可显示的历史记录。")

    if delete_mode:
        action_col1, action_col2 = st.columns(2, gap="small")
        with action_col1:
            delete_clicked = st.button(
                "删除所选记录",
                width="stretch",
                disabled=not selected_ids,
            )
        with action_col2:
            cancel_clicked = st.button(
                "取消删除",
                width="stretch",
            )
    else:
        delete_clicked = st.button(
            "删除记录",
            width="stretch",
            disabled=not history,
        )
        cancel_clicked = False

    if not delete_mode and delete_clicked:
        st.session_state["history_delete_mode"] = True
        st.rerun()

    if delete_mode and cancel_clicked:
        st.session_state["history_delete_mode"] = False
        st.rerun()

    if clear_search_clicked:
        st.session_state["history_clear_search_requested"] = True
        st.rerun()

    if delete_mode and delete_clicked and selected_ids:
        selected_set = set(selected_ids)
        new_history = [item for item in history if str(item.get("id", "")) not in selected_set]
        st.session_state["translation_history"] = new_history
        st.session_state["history_delete_mode"] = False
        try:
            _save_translation_history(new_history)
            _notify_success(f"已删除 {len(history) - len(new_history)} 条历史记录。")
        except Exception as exc:
            st.warning(f"删除历史记录失败: {exc}")
        st.rerun()


def _render_translation_corpus() -> None:
    st.checkbox(
        "翻译时使用语料库",
        key="corpus_enabled",
        help="开启后，已归档的参考内容会作为翻译规则的一部分发送给模型，用于统一术语和表达风格。",
    )
    uploaded_refs = st.file_uploader(
        "导入翻译参考文件",
        type=CORPUS_UPLOAD_TYPES,
        accept_multiple_files=True,
        key="corpus_uploads",
        help="支持 PDF、docx、pptx、xlsx、txt、md、csv。旧版 doc/ppt/xls 请先另存为新版格式。",
    )
    if st.button(
        "归档到语料库",
        key="import_corpus_refs",
        width="stretch",
        disabled=not uploaded_refs,
    ):
        entries = list(st.session_state.get("translation_corpus", []))
        memory = st.session_state.get("translation_memory", _empty_translation_memory())
        if not isinstance(memory, dict):
            memory = _empty_translation_memory()
        imported = 0
        imported_terms = 0
        failed: list[str] = []
        for uploaded in uploaded_refs or []:
            try:
                text = _extract_reference_text(uploaded.name, uploaded.getvalue())
                if not text:
                    failed.append(f"{uploaded.name}: 未提取到可用文本")
                    continue
                entries = [entry for entry in entries if entry.get("filename") != uploaded.name]
                entries.append(
                    {
                        "id": str(uuid.uuid4()),
                        "filename": uploaded.name,
                        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "text": text,
                    }
                )
                addition = _extract_memory_from_uploaded_reference(uploaded.name, uploaded.getvalue(), text)
                imported_terms += len(addition.get("terms", []))
                memory = _merge_translation_memory(memory, addition)
                imported += 1
            except Exception as exc:
                failed.append(f"{uploaded.name}: {exc}")

        st.session_state["translation_corpus"] = entries
        st.session_state["translation_memory"] = memory
        try:
            _save_translation_corpus(entries)
            _save_translation_memory(memory)
            if imported:
                _notify_success(f"已归档 {imported} 个参考文件，并更新 {imported_terms} 条术语记忆。")
        except Exception as exc:
            st.warning(f"保存语料库失败: {exc}")
        for message in failed[:5]:
            st.warning(message)

    entries = st.session_state.get("translation_corpus", [])
    memory = st.session_state.get("translation_memory", _empty_translation_memory())
    if not isinstance(memory, dict):
        memory = _empty_translation_memory()
    term_count = len(memory.get("terms", [])) if isinstance(memory.get("terms", []), list) else 0
    pending_count = (
        sum(1 for item in memory.get("terms", []) if isinstance(item, dict) and not str(item.get("target", "")).strip())
        if isinstance(memory.get("terms", []), list)
        else 0
    )
    example_count = len(memory.get("examples", [])) if isinstance(memory.get("examples", []), list) else 0
    st.caption(
        f"当前记忆：{term_count} 条术语，{pending_count} 条待补全中文，"
        f"{example_count} 条风格参考。翻译时只按相关性调用，不注入完整原文。"
    )

    terms = memory.get("terms", [])
    if isinstance(terms, list) and terms:
        term_rows = [
            {
                "英文/缩写": str(item.get("source", "")),
                "中文译法": str(item.get("target", "")),
                "完整拼写": str(item.get("full_form", "")),
                "来源文件": str(item.get("source_file", "")),
            }
            for item in terms
            if isinstance(item, dict)
        ]
        edited_terms = st.data_editor(
            pd.DataFrame(term_rows),
            width="stretch",
            hide_index=True,
            key="translation_memory_terms_editor",
            column_config={
                "中文译法": st.column_config.TextColumn("中文译法", help="可直接修改术语译法"),
            },
            disabled=["英文/缩写", "完整拼写", "来源文件"],
        )
        edit_col1, edit_col2 = st.columns(2, gap="small")
        with edit_col1:
            save_terms_clicked = st.button("保存术语修改", key="save_memory_terms", width="stretch")
        with edit_col2:
            complete_terms_clicked = st.button(
                "自动补全缺失中文",
                key="complete_missing_memory_terms",
                width="stretch",
                disabled=pending_count == 0,
            )
        if save_terms_clicked:
            edited_records = edited_terms.to_dict("records")
            for idx, row in enumerate(edited_records):
                if idx < len(terms) and isinstance(terms[idx], dict):
                    terms[idx]["target"] = str(row.get("中文译法", "")).strip()
            memory["terms"] = terms
            memory["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state["translation_memory"] = memory
            try:
                _save_translation_memory(memory)
                _notify_success("术语修改已保存。")
            except Exception as exc:
                st.warning(f"保存术语修改失败: {exc}")
            st.rerun()
        if complete_terms_clicked:
            active_group = st.session_state.get("runtime_provider_group", "API Providers")
            active_config = (
                st.session_state.get("api_provider_config", _default_api_provider_config())
                if active_group == "API Providers"
                else st.session_state.get("local_provider_config", _default_local_provider_config())
            )
            validation_error = _validate_runtime_config(active_config, active_group)
            if validation_error:
                st.warning(validation_error)
            else:
                try:
                    with st.spinner("正在按建筑/施工图/建筑可视化语境补全缺失中文..."):
                        completed = _complete_missing_memory_targets(memory, active_config)
                    st.session_state["translation_memory"] = memory
                    _save_translation_memory(memory)
                    _notify_success(f"已补全 {completed} 条术语中文。")
                    st.rerun()
                except Exception as exc:
                    st.warning(f"自动补全失败: {exc}")

    if isinstance(entries, list) and entries:
        rows = [
            {
                "文件名": entry.get("filename", ""),
                "录入时间": entry.get("created_at", ""),
                "字符数": len(entry.get("text", "")),
            }
            for entry in entries
        ]
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
        if st.button("从语料库重建记忆", key="rebuild_translation_memory", width="stretch"):
            rebuilt = _rebuild_memory_from_corpus(entries)
            st.session_state["translation_memory"] = rebuilt
            try:
                _save_translation_memory(rebuilt)
                _notify_success("语料库记忆已重建。")
            except Exception as exc:
                st.warning(f"重建记忆失败: {exc}")
            st.rerun()
        if st.button("清空语料库", key="clear_translation_corpus", width="stretch"):
            st.session_state["translation_corpus"] = []
            st.session_state["translation_memory"] = _empty_translation_memory()
            try:
                _save_translation_corpus([])
                _save_translation_memory(_empty_translation_memory())
                _notify_success("语料库已清空。")
            except Exception as exc:
                st.warning(f"清空语料库失败: {exc}")
            st.rerun()
    else:
        st.caption("语料库为空。导入参考文件后，后续翻译会自动调用其中的术语和表达习惯。")


def main() -> None:
    st.set_page_config(page_title=f"{APP_NAME} Studio", layout="wide")
    _inject_styles()
    _init_state()

    st.markdown(
        f"""
<div class="hero">
    <h1>{APP_NAME} <span style="font-weight:500;opacity:0.88;">{APP_VERSION}</span></h1>
    <p>Upload multiple PDFs, preview detected comments, estimate token cost, and translate in one batch.</p>
</div>
        """,
        unsafe_allow_html=True,
    )

    source_labels = [label for label, _ in SOURCE_LANGUAGE_OPTIONS]
    target_labels = [label for label, _ in LANGUAGE_OPTIONS]
    source_map = {label: code for label, code in SOURCE_LANGUAGE_OPTIONS}
    target_map = {label: code for label, code in LANGUAGE_OPTIONS}
    current_group = st.session_state.get("runtime_provider_group", "API Providers")
    current_active_config = (
        st.session_state.get("api_provider_config", _default_api_provider_config())
        if current_group == "API Providers"
        else st.session_state.get("local_provider_config", _default_local_provider_config())
    )
    vision_supported, vision_reason = _is_vision_supported(current_active_config, current_group)

    progress_col, stop_col = st.columns([7.4, 1.0], gap="small")
    with progress_col:
        progress_bar = st.empty()
        _render_stage_progress(progress_bar)
    with stop_col:
        st.markdown('<div class="stop-button-spacer"></div>', unsafe_allow_html=True)
        top_stop_clicked = st.button(
            "终止",
            key="stop_translation_btn",
            type="primary" if bool(st.session_state.get("translation_running", False)) else "secondary",
            width="stretch",
            disabled=not bool(st.session_state.get("translation_running", False)),
            help="终止当前翻译任务。正在等待单次本地模型推理时，会在当前请求返回或超时后停止后续流程。",
        )
    if top_stop_clicked:
        _request_translation_cancel()
        _render_stage_progress(progress_bar)

    left_col, right_col = st.columns([1.25, 1], gap="large")
    with left_col:
        uploaded_pdfs = st.file_uploader(
            "1) Upload PDFs",
            type=["pdf"],
            accept_multiple_files=True,
        )
        current_upload_signature = _uploaded_files_signature(uploaded_pdfs)
        previous_upload_signature = str(st.session_state.get("uploaded_files_signature", ""))
        if current_upload_signature != previous_upload_signature:
            _reset_run_state_for_upload_change()
            st.session_state["uploaded_files_signature"] = current_upload_signature
            _render_stage_progress(progress_bar)
        pdf_type_reports = _classify_uploaded_pdfs(uploaded_pdfs) if uploaded_pdfs else []
        recommended_mode = _render_pdf_type_reports(
            pdf_type_reports,
            vision_supported=vision_supported,
        ) if uploaded_pdfs else "hybrid"
        source_lang_label = st.selectbox("2) Source language", options=source_labels, index=0)
        target_lang_label = st.selectbox("3) Target language", options=target_labels, index=0)
        mode_title_col, mode_refresh_col = st.columns([5, 1], gap="small")
        with mode_title_col:
            st.caption("4) Mode")
        with mode_refresh_col:
            mode_refresh_clicked = st.button(
                "刷新检测",
                key="refresh_mode_detection",
                width="stretch",
                help="重新检测当前配置是否支持视觉模式，并刷新预估结果。",
            )
        if mode_refresh_clicked:
            _refresh_detection_cache()
            st.session_state["preflight_reports"] = []
            st.session_state["preflight_key"] = ""
            st.rerun()
        mode_options = list(MODE_HINTS.keys()) if vision_supported else ["text extraction only"]
        if recommended_mode not in mode_options:
            recommended_mode = "text extraction only" if "text extraction only" in mode_options else mode_options[0]
        mode_label = st.radio(
            "Mode",
            options=mode_options,
            horizontal=True,
            index=mode_options.index(recommended_mode),
        )
        if not vision_supported:
            st.info(
                f"当前生效配置（{current_group} / {current_active_config.get('provider_name', '')}）"
                f"不支持视觉模式，已自动限制为 text extraction only。{vision_reason}"
            )
        elif vision_reason:
            st.caption(vision_reason)
        st.caption(MODE_HINTS[mode_label])
        if _server_path_saving_enabled():
            st.text_input(
                "自动保存路径",
                key="export_output_dir",
                help="翻译完成后会自动保存翻译后的 PDF 到该目录。",
            )
        else:
            st.info("网页云端运行时不会使用服务器保存路径；翻译完成后请通过浏览器下载结果。")
        action_col1, action_col2 = st.columns(2, gap="small")
        with action_col1:
            analyze_clicked = st.button(
                "Analyze / Estimate",
                width="stretch",
                disabled=not uploaded_pdfs,
                help="按需执行预分析（提取预览 + token/cost估算）。",
            )
        with action_col2:
            translate_clicked = st.button(
                "Translate",
                type="primary",
                width="stretch",
                disabled=not uploaded_pdfs,
            )
        fast_mode_enabled = st.checkbox(
            "极速模式",
            key="fast_mode_enabled",
            help="速度优先：降低视觉请求负载并使用更激进的超时/重试参数，以提升整体翻译吞吐。",
        )
        if fast_mode_enabled:
            st.caption("开启后通常会更快，但极小手写体的识别稳定性可能略有下降。")
            fm_col1, fm_col2 = st.columns(2, gap="small")
            with fm_col1:
                fast_mode_dpi = st.slider(
                    "极速 Vision DPI",
                    min_value=96,
                    max_value=220,
                    step=2,
                    key="fast_mode_dpi",
                )
                fast_mode_timeout_seconds = st.slider(
                    "视觉超时(秒)",
                    min_value=60,
                    max_value=360,
                    step=10,
                    key="fast_mode_timeout_seconds",
                )
            with fm_col2:
                fast_mode_max_tokens = st.slider(
                    "视觉最大输出Tokens",
                    min_value=512,
                    max_value=8192,
                    step=256,
                    key="fast_mode_max_tokens",
                )
                fast_mode_retry_attempts = st.slider(
                    "视觉重试次数",
                    min_value=1,
                    max_value=3,
                    step=1,
                    key="fast_mode_retry_attempts",
                )
        else:
            fast_mode_dpi = int(st.session_state.get("fast_mode_dpi", 150))
            fast_mode_timeout_seconds = int(st.session_state.get("fast_mode_timeout_seconds", 180))
            fast_mode_max_tokens = int(st.session_state.get("fast_mode_max_tokens", 2048))
            fast_mode_retry_attempts = int(st.session_state.get("fast_mode_retry_attempts", 2))

        odl_fallback_enabled = st.checkbox(
            "启用 ODL 结构化回退（实验）",
            key="odl_fallback_enabled",
            help=(
                "默认关闭。仅在疑似文本切割异常的页面触发 opendataloader-pdf 回退提取。"
                "未安装 opendataloader-pdf 或 Java 时会自动忽略，不影响主流程。"
            ),
        )
        if odl_fallback_enabled:
            st.caption("建议仅用于复杂大画布（如 Miro 导出），以减少额外开销。")

        with st.expander("Cost estimation assumptions", expanded=False):
            usd_to_cny_rate = st.number_input(
                "USD -> CNY exchange rate",
                min_value=0.01,
                step=0.01,
                key="usd_to_cny_rate",
            )
            prompt_overhead_tokens = st.number_input(
                "Prompt overhead tokens per text comment",
                min_value=1,
                step=1,
                key="prompt_overhead_tokens",
            )
            vision_input_tokens_per_page = st.number_input(
                "Vision input tokens per page",
                min_value=1,
                step=1,
                key="vision_input_tokens_per_page",
            )
            vision_output_tokens_per_page = st.number_input(
                "Vision output tokens per page",
                min_value=1,
                step=1,
                key="vision_output_tokens_per_page",
            )
            input_cost_per_1m = st.number_input(
                "Input cost per 1M tokens (USD)",
                min_value=0.0,
                step=0.05,
                key="input_cost_per_1m",
            )
            output_cost_per_1m = st.number_input(
                "Output cost per 1M tokens (USD)",
                min_value=0.0,
                step=0.05,
                key="output_cost_per_1m",
            )

    with right_col:
        advanced_title_col, advanced_save_col = st.columns([2, 1], gap="small")
        with advanced_title_col:
            st.subheader("Advanced")
        with advanced_save_col:
            st.write("")
            if st.button("保存设置", key="save_all_advanced_settings", width="stretch", type="primary"):
                _persist_ui_settings()
                _refresh_detection_cache()
                st.session_state["preflight_reports"] = []
                st.session_state["preflight_key"] = ""
                _notify_success("设置已保存。")
        with st.container(height=720, border=False):
            st.caption("Provider Settings")
            previous_group = st.session_state.get("runtime_provider_group", "API Providers")
            st.radio(
                "Use configuration from",
                options=["API Providers", "Local Models"],
                horizontal=True,
                key="runtime_provider_group",
            )
            current_group_after_input = st.session_state.get("runtime_provider_group", "API Providers")
            if current_group_after_input != previous_group:
                _refresh_detection_cache()
                st.rerun()

            if current_group_after_input == "API Providers":
                st.markdown("**API Providers**")
                api_names = list(API_PROVIDER_PRESETS.keys())
                st.selectbox(
                    "Providers",
                    options=api_names,
                    key="api_provider_name_ui",
                    on_change=_on_api_provider_name_change,
                )
                st.text_input("Base URL", key="api_base_url_ui")
                st.text_input("Model", key="api_model_ui")
                st.text_input("API key", key="api_api_key_ui", type="password")
                st.text_area(
                    "翻译设定（Prompts）",
                    key="api_prompt_ui",
                    height=110,
                    help="可写入固定翻译规则，例如术语表、行业表达偏好、禁止逐字翻译等。",
                )

                api_config = _build_config_from_inputs(
                    API_PROVIDER_PRESETS,
                    str(st.session_state.get("api_provider_name_ui", "OpenAI")),
                    str(st.session_state.get("api_base_url_ui", "")),
                    str(st.session_state.get("api_api_key_ui", "")),
                    str(st.session_state.get("api_model_ui", "")),
                    str(st.session_state.get("api_prompt_ui", "")),
                )
                st.session_state["api_provider_config"] = api_config

                if st.button("Check connection", key="check_api_connection_btn", width="stretch"):
                    ok, message = _check_connection(api_config)
                    if ok:
                        _notify_success(message)
                    else:
                        _notify_error(message)
            else:
                st.markdown("**Local Models**")
                local_names = list(LOCAL_MODEL_PRESETS.keys())
                st.selectbox(
                    "Providers",
                    options=local_names,
                    key="local_provider_name_ui",
                    on_change=_on_local_provider_name_change,
                )
                st.text_input("Base URL", key="local_base_url_ui")
                st.text_input("Model", key="local_model_ui")
                st.text_input("API key", key="local_api_key_ui", type="password")
                st.text_area(
                    "翻译设定（Prompts）",
                    key="local_prompt_ui",
                    height=110,
                    help="可写入固定翻译规则，例如术语表、行业表达偏好、禁止逐字翻译等。",
                )

                local_config = _build_config_from_inputs(
                    LOCAL_MODEL_PRESETS,
                    str(st.session_state.get("local_provider_name_ui", "Ollama")),
                    str(st.session_state.get("local_base_url_ui", "")),
                    str(st.session_state.get("local_api_key_ui", "")),
                    str(st.session_state.get("local_model_ui", "")),
                    str(st.session_state.get("local_prompt_ui", "")),
                )
                st.session_state["local_provider_config"] = local_config

                if st.button("Check connection", key="check_local_connection_btn", width="stretch"):
                    ok, message = _check_connection(local_config)
                    if ok:
                        _notify_success(message)
                    else:
                        _notify_error(message)

            with st.expander("语料库", expanded=False):
                _render_translation_corpus()

            st.markdown("---")
            typed_text_threshold = st.slider(
                "Hybrid text threshold",
                min_value=1,
                max_value=400,
                key="typed_text_threshold",
                help=(
                    "仅在 hybrid 模式生效：页面文本字符数低于该阈值时，更倾向使用视觉识别。\n"
                    "简单理解：调大更容易触发视觉（更稳识别手写，但更慢更贵）；"
                    "调小更偏向文本（更快更省，但手写页可能漏）。"
                ),
            )
            dpi_for_vision = st.slider("Vision DPI", min_value=72, max_value=400, key="dpi_for_vision")
            font_size = st.selectbox(
                "Annotation font size",
                options=ADOBE_ANNOTATION_FONT_SIZES,
                key="font_size",
                format_func=lambda x: str(int(x)),
                help="Matches Bluebeam annotation font size range: 2-72.",
            )
            placement_label = st.radio("Placement", options=list(PLACEMENT_OPTIONS.keys()), key="placement_label")
            max_pages_raw = st.text_input("Max pages per PDF (optional)", key="max_pages_raw")

    preflight_reports: list[dict[str, Any]] = []
    if uploaded_pdfs:
        preflight_key = _build_preflight_key(
            uploaded_pdfs,
            mode=mode_label,
            typed_text_threshold=int(typed_text_threshold),
            prompt_overhead_tokens=int(prompt_overhead_tokens),
            vision_input_tokens_per_page=int(vision_input_tokens_per_page),
            vision_output_tokens_per_page=int(vision_output_tokens_per_page),
            input_cost_per_1m=float(input_cost_per_1m),
            output_cost_per_1m=float(output_cost_per_1m),
        )

        if analyze_clicked:
            with st.spinner("Analyzing uploaded PDFs..."):
                preflight_reports = []
                for uploaded in uploaded_pdfs:
                    report = _analyze_uploaded_pdf(
                        pdf_bytes=uploaded.getvalue(),
                        filename=uploaded.name,
                        mode=mode_label,
                        typed_text_threshold=int(typed_text_threshold),
                        prompt_overhead_tokens=int(prompt_overhead_tokens),
                        vision_input_tokens_per_page=int(vision_input_tokens_per_page),
                        vision_output_tokens_per_page=int(vision_output_tokens_per_page),
                        input_cost_per_1m=float(input_cost_per_1m),
                        output_cost_per_1m=float(output_cost_per_1m),
                    )
                    preflight_reports.append(report)
            st.session_state["preflight_reports"] = preflight_reports
            st.session_state["preflight_key"] = preflight_key
        elif st.session_state.get("preflight_key", "") == preflight_key:
            cached_reports = st.session_state.get("preflight_reports", [])
            if isinstance(cached_reports, list):
                preflight_reports = cached_reports
        else:
            st.session_state["preflight_reports"] = []
            st.session_state["preflight_key"] = ""

        if preflight_reports:
            _render_preflight_reports(preflight_reports, float(st.session_state["usd_to_cny_rate"]))
        else:
            st.caption("点击 Analyze / Estimate 生成提取预览与消耗预测。")
    else:
        st.session_state["preflight_reports"] = []
        st.session_state["preflight_key"] = ""

    if translate_clicked:
        if not uploaded_pdfs:
            st.error("Please upload at least one PDF.")
            return

        if fast_mode_enabled:
            os.environ["AUTOPDFTRANSLATOR_VISION_TIMEOUT_SECONDS"] = str(int(fast_mode_timeout_seconds))
            os.environ["AUTOPDFTRANSLATOR_VISION_MAX_TOKENS"] = str(int(fast_mode_max_tokens))
            os.environ["AUTOPDFTRANSLATOR_VISION_RETRY_ATTEMPTS"] = str(int(fast_mode_retry_attempts))
            effective_dpi_for_vision = int(fast_mode_dpi)
        else:
            os.environ.pop("AUTOPDFTRANSLATOR_VISION_TIMEOUT_SECONDS", None)
            os.environ.pop("AUTOPDFTRANSLATOR_VISION_MAX_TOKENS", None)
            os.environ.pop("AUTOPDFTRANSLATOR_VISION_RETRY_ATTEMPTS", None)
            effective_dpi_for_vision = int(dpi_for_vision)
        if _is_mode_requires_vision(mode_label) and any(
            str(report.get("pdf_type", "")) in {"图片/手写混合型 PDF", "扫描/图片型 PDF"}
            for report in pdf_type_reports
        ):
            effective_dpi_for_vision = max(effective_dpi_for_vision, 300)

        debug_root = Path(str(st.session_state.get("export_output_dir", "") or "").strip())
        if debug_root:
            os.environ["AUTOPDFTRANSLATOR_VISION_DEBUG_DIR"] = str(debug_root / "_vision_debug")
        else:
            os.environ.pop("AUTOPDFTRANSLATOR_VISION_DEBUG_DIR", None)

        if bool(st.session_state.get("odl_fallback_enabled", False)):
            os.environ["AUTOPDFTRANSLATOR_ODL_FALLBACK"] = "1"
        else:
            os.environ.pop("AUTOPDFTRANSLATOR_ODL_FALLBACK", None)

        _persist_ui_settings()

        active_group = st.session_state["runtime_provider_group"]
        active_config = (
            st.session_state["api_provider_config"]
            if active_group == "API Providers"
            else st.session_state["local_provider_config"]
        )
        validation_error = _validate_runtime_config(active_config, active_group)
        if validation_error:
            st.error(validation_error)
            return

        if _is_mode_requires_vision(mode_label):
            vis_ok, vis_reason = _is_vision_supported(active_config, active_group)
            if not vis_ok:
                st.error(f"当前配置不支持视觉模式：{vis_reason} 请切换到 text extraction only 或更换视觉模型。")
                return

        provider_runtime = _apply_runtime_provider(active_config)

        max_pages = None
        if max_pages_raw.strip():
            try:
                max_pages = max(1, int(max_pages_raw))
            except ValueError:
                st.error("Max pages must be an integer.")
                return

        artifacts: list[dict[str, Any]] = []
        st.session_state["translation_running"] = True
        st.session_state["translation_cancel_requested"] = False
        st.session_state["task_progress_percent"] = 0
        st.session_state["task_progress_text"] = "翻译中-进度：0%"
        st.session_state["task_progress_stage"] = "extracting"
        _render_stage_progress(progress_bar)
        placement = PLACEMENT_OPTIONS[placement_label]
        source_lang = source_map[source_lang_label]
        target_lang = target_map[target_lang_label]
        task_started_at = time.perf_counter()

        try:
            for idx, uploaded in enumerate(uploaded_pdfs, start=1):
                _raise_if_translation_cancelled()
                pct = int((idx - 1) / len(uploaded_pdfs) * 100)
                progress_text = f"翻译中-进度：{pct}%"
                st.session_state["task_progress_percent"] = pct
                st.session_state["task_progress_text"] = progress_text
                st.session_state["task_progress_stage"] = "extracting"
                _render_stage_progress(progress_bar)

                log_handler = _StreamlitLogHandler()
                log_handler.setFormatter(logging.Formatter("%(levelname)s %(message)s"))
                root_logger = logging.getLogger()
                previous_level = root_logger.level
                root_logger.addHandler(log_handler)
                root_logger.setLevel(logging.INFO)

                try:
                    with tempfile.TemporaryDirectory() as temp_dir:
                        temp_root = Path(temp_dir)
                        input_path = temp_root / uploaded.name
                        output_path = temp_root / f"{input_path.stem}.translated.pdf"
                        input_path.write_bytes(uploaded.getvalue())

                        results = translate_pdf(
                            input_pdf=input_path,
                            output_pdf=output_path,
                            source_lang=source_lang,
                            target_lang=target_lang,
                            mode=mode_label,
                            provider=provider_runtime,
                            typed_text_threshold=int(typed_text_threshold),
                            dpi_for_vision=int(effective_dpi_for_vision),
                            dry_run=False,
                            max_pages=max_pages,
                            layout=LayoutConfig(
                                font_size=float(font_size),
                                placement=placement,
                            ),
                            progress_callback=_build_task_progress_callback(
                                progress_bar,
                                idx,
                                len(uploaded_pdfs),
                            ),
                        )

                        artifacts.append(
                            {
                                "filename": uploaded.name,
                                "output_pdf_name": output_path.name,
                                "output_pdf_bytes": output_path.read_bytes(),
                                "results_json_bytes": _results_to_json(results),
                                "page_stats": _build_page_stats(results),
                                "log_lines": list(log_handler.lines),
                            }
                        )
                finally:
                    root_logger.removeHandler(log_handler)
                    root_logger.setLevel(previous_level)

            zip_bytes, zip_name = _build_zip(artifacts)
            st.session_state["run_error"] = ""
            total_pages_done = sum(len(item["page_stats"]) for item in artifacts)
            total_vision_pages_done = sum(
                sum(1 for row in item["page_stats"] if row["Vision used"])
                for item in artifacts
            )
            total_text_pages_done = max(0, total_pages_done - total_vision_pages_done)
            base_message = (
                f"Translation complete: {len(artifacts)} PDF(s) finished. "
                f"Vision pages: {total_vision_pages_done}/{total_pages_done}; "
                f"Text pages: {total_text_pages_done}/{total_pages_done}."
            )
            st.session_state["artifacts"] = artifacts
            st.session_state["batch_zip_bytes"] = zip_bytes
            st.session_state["batch_zip_name"] = zip_name

            if _server_path_saving_enabled():
                auto_ok, auto_message = _save_artifacts_to_path(
                    artifacts,
                    str(st.session_state.get("export_output_dir", "")),
                    zip_bytes,
                    zip_name,
                    save_json=False,
                    save_zip=False,
                )
            else:
                auto_ok = False
                auto_message = "云端运行不保存到服务器路径；请使用浏览器下载结果。"
            st.session_state["save_path_message"] = auto_message
            if auto_ok:
                _persist_ui_settings()
                st.session_state["run_message"] = f"{base_message} {auto_message}"
            else:
                st.session_state["run_message"] = f"{base_message} {auto_message}"

            failure_reason = _artifact_failure_reason(artifacts, mode_label)
            history_entries = _build_history_entries(
                artifacts,
                export_output_dir=str(st.session_state.get("export_output_dir", "")),
                mode_label=mode_label,
                provider_group=active_group,
                provider_config=active_config,
                auto_saved=auto_ok,
                task_duration_seconds=(time.perf_counter() - task_started_at),
            )
            _append_translation_history(history_entries)

            st.session_state["task_progress_percent"] = 100
            if failure_reason:
                st.session_state["run_error"] = f"任务失败（{failure_reason}），已保留当前 PDF 结果供检查。"
                st.session_state["run_message"] = ""
                st.session_state["task_progress_text"] = f"任务失败：{failure_reason}"
                st.session_state["task_progress_stage"] = "failed"
                st.session_state["failure_animation_run_id"] = str(uuid.uuid4())
            else:
                done_text = "任务已完成，请从下方查看翻译结果"
                st.session_state["task_progress_text"] = done_text
                st.session_state["task_progress_stage"] = "done"
                st.session_state["confetti_run_id"] = str(uuid.uuid4())
                st.session_state["browser_download_run_id"] = str(uuid.uuid4())
            st.session_state["translation_running"] = False
            st.session_state["translation_cancel_requested"] = False
            _render_stage_progress(progress_bar)
        except TranslationCancelled as exc:
            st.session_state["run_error"] = ""
            st.session_state["translation_running"] = False
            st.session_state["translation_cancel_requested"] = False
            if artifacts:
                zip_bytes, zip_name = _build_zip(artifacts)
                st.session_state["artifacts"] = artifacts
                st.session_state["batch_zip_bytes"] = zip_bytes
                st.session_state["batch_zip_name"] = zip_name
                st.session_state["run_message"] = (
                    f"任务已终止，已保留 {len(artifacts)} 个已完成 PDF。"
                    "请从下方下载当前结果。"
                )
            else:
                st.session_state["run_message"] = "任务已终止，未生成翻译结果。"
                st.session_state["artifacts"] = []
                st.session_state["batch_zip_bytes"] = b""
                st.session_state["batch_zip_name"] = ""
            st.session_state["task_progress_text"] = str(exc)
            st.session_state["task_progress_stage"] = "idle"
            _render_stage_progress(progress_bar)
        except Exception as exc:
            st.session_state["translation_running"] = False
            st.session_state["translation_cancel_requested"] = False
            st.session_state["run_error"] = str(exc)
            if artifacts:
                zip_bytes, zip_name = _build_zip(artifacts)
                st.session_state["artifacts"] = artifacts
                st.session_state["batch_zip_bytes"] = zip_bytes
                st.session_state["batch_zip_name"] = zip_name
                st.session_state["run_message"] = (
                    f"翻译中断，但已保留 {len(artifacts)} 个已完成 PDF。"
                    "请从下方下载当前结果。"
                )
                partial_pct = max(1, min(99, int(len(artifacts) / len(uploaded_pdfs) * 100)))
                st.session_state["task_progress_percent"] = partial_pct
                st.session_state["task_progress_text"] = "翻译中断，已保留当前结果"
                st.session_state["task_progress_stage"] = "failed"
                st.session_state["failure_animation_run_id"] = str(uuid.uuid4())
                _render_stage_progress(progress_bar)
            else:
                st.session_state["run_message"] = ""
                st.session_state["artifacts"] = []
                st.session_state["batch_zip_bytes"] = b""
                st.session_state["batch_zip_name"] = ""
                st.session_state["task_progress_percent"] = 0
                st.session_state["task_progress_text"] = "翻译失败，请检查日志后重试"
                st.session_state["task_progress_stage"] = "failed"
                st.session_state["failure_animation_run_id"] = str(uuid.uuid4())
                _render_stage_progress(progress_bar)

    _render_outputs()
    _render_translation_history()


if __name__ == "__main__":
    main()
